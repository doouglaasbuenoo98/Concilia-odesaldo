from http.server import HTTPServer, SimpleHTTPRequestHandler
import subprocess, json, os, threading, webbrowser, io

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
PORT       = 5000


def _gerar_xlsx(headers: list, rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COR_HEADER   = "1F4E79"
    COR_ZEBRA    = "DEEAF1"
    COR_PENDENTE = "FFDDC1"
    COR_PARCIAL  = "FFF2CC"
    COR_AJUSTADO = "C6EFCE"

    # Formato contábil brasileiro: R$ com separador de milhar e 2 casas decimais
    FMT_BRL   = '_-"R$"* #,##0.00_-;-"R$"* #,##0.00_-;_-"R$"* "-"??_-;_-@_-'
    # Formato quantidade: até 3 casas decimais, sem moeda
    FMT_QTY   = '#,##0.###'

    # Colunas que recebem formato de moeda R$
    COLS_BRL  = {"Valor Div. (R$)", "Valor Ajustado (R$)"}
    # Colunas que recebem formato de quantidade
    COLS_QTY  = {"Qtd WMS", "Qtd ERP", "Diff Original", "Diff Restante", "Qtd Ajuste"}

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def to_number(val):
        """Converte string (formato BR ou EN) para número.
        Retorna int quando o valor é inteiro (evita vírgula pendente no Excel).
        """
        if val is None or val == "":
            return None
        s = str(val).strip()
        try:
            if "," in s and "." in s:
                # formato BR: ponto=milhar, vírgula=decimal  ex: "10.316,08"
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                # só vírgula → decimal  ex: "-757,92"
                s = s.replace(",", ".")
            # else: ponto como decimal EN  ex: "-757.92" ou inteiro puro
            f = float(s)
            return int(f) if f == int(f) else f
        except (ValueError, TypeError):
            return val

    wb = Workbook()
    ws = wb.active
    ws.title = "Divergencias"

    # Mapear header → formato de número (col_idx 1-based → format_str)
    col_fmt = {}
    for col_idx, h in enumerate(headers, 1):
        if h in COLS_BRL:
            col_fmt[col_idx] = FMT_BRL
        elif h in COLS_QTY:
            col_fmt[col_idx] = FMT_QTY

    # Cabeçalho
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = fill(COR_HEADER)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border()
    ws.row_dimensions[1].height = 22

    # Encontrar índice da coluna Status para colorir linhas
    try:
        idx_status = headers.index("Status")
    except ValueError:
        idx_status = -1

    for row_num, row_data in enumerate(rows, 2):
        status_val = row_data[idx_status] if idx_status >= 0 and idx_status < len(row_data) else ""
        if status_val == "Ajustado":
            row_fill = fill(COR_AJUSTADO)
        elif status_val == "Parcial":
            row_fill = fill(COR_PARCIAL)
        elif status_val == "Pendente":
            row_fill = fill(COR_PENDENTE)
        else:
            row_fill = fill(COR_ZEBRA) if row_num % 2 == 0 else None

        for col_idx, val in enumerate(row_data, 1):
            fmt = col_fmt.get(col_idx)
            cell_val = to_number(val) if fmt else val
            cell = ws.cell(row=row_num, column=col_idx, value=cell_val)
            if fmt:
                # Inteiros não precisam de formato decimal (evita vírgula pendente no Excel BR)
                if fmt == FMT_QTY and isinstance(cell_val, int):
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = fmt
                cell.alignment = Alignment(vertical="center", horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill
            cell.border = border()
            cell.font   = Font(size=10)

    # Auto-filter e freeze
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes    = "A2"

    # Ajustar largura das colunas
    col_widths = {h: max(len(h), 8) for h in headers}
    for row_data in rows:
        for h, val in zip(headers, row_data):
            col_widths[h] = min(max(col_widths[h], len(str(val or "")) + 2), 45)
    # Colunas de moeda merecem espaço extra para o R$
    for h in COLS_BRL:
        if h in col_widths:
            col_widths[h] = max(col_widths[h], 18)
    for col_idx, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[h]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=OUTPUT_DIR, **kwargs)

    def do_POST(self):
        if self.path == "/api/atualizar":
            result = subprocess.run(
                ["python", "reconciliacao.py"],
                capture_output=True, text=True, cwd=BASE_DIR
            )
            ok   = result.returncode == 0
            msg  = result.stdout if ok else result.stderr
            body = json.dumps({"ok": ok, "msg": msg}).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", len(body))
            self.end_headers()
            self.wfile.write(body)

        elif self.path == "/api/atualizar_saida":
            saida_script = os.path.join(os.path.dirname(BASE_DIR), "Saida", "reconciliacao.py")
            result = subprocess.run(
                ["python", saida_script],
                capture_output=True, text=True, cwd=BASE_DIR
            )
            ok   = result.returncode == 0
            msg  = result.stdout if ok else result.stderr
            body = json.dumps({"ok": ok, "msg": msg}).encode("utf-8")
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", len(body))
            self.end_headers()
            self.wfile.write(body)

        elif self.path == "/api/exportar_divergencias":
            content_length = int(self.headers.get("Content-Length", 0))
            body_raw = self.rfile.read(content_length)
            data = json.loads(body_raw.decode("utf-8"))
            xlsx = _gerar_xlsx(data["headers"], data["rows"])
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="divergencias.xlsx"')
            self.send_header("Content-Length", len(xlsx))
            self.end_headers()
            self.wfile.write(xlsx)

        else:
            self.send_error(404)

    def log_message(self, format, *args):
        pass


if __name__ == "__main__":
    import sys
    abrir_browser = "--no-browser" not in sys.argv
    try:
        server = HTTPServer(("localhost", PORT), Handler)
    except OSError:
        sys.exit(0)
    url = f"http://localhost:{PORT}/index.html"
    if abrir_browser:
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    server.serve_forever()
