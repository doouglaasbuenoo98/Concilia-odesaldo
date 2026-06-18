import pandas as pd
import sys
import os
import re
import json
from datetime import datetime
from config import WMS_COLUNAS, ERP_COLUNAS, TOLERANCIA_DIAS, CHAVE_CONCILIACAO, WMS_FILTRO_STATUS, ERP_SKU_PADRAO

OUTPUT_DIR      = "output"
DADOS_DIR       = "dados"
TEMPLATE_PATH   = os.path.join("templates", "dashboard.html")
HISTORICO_EXCEL = os.path.join(OUTPUT_DIR, "historico.xlsx")
HISTORICO_HTML  = os.path.join(OUTPUT_DIR, "historico.html")
HOME_HTML       = os.path.join(OUTPUT_DIR, "index.html")


# ── Carregamento ──────────────────────────────────────────────────────────────

def carregar_arquivo(caminho: str, mapeamento: dict) -> pd.DataFrame:
    ext = os.path.splitext(caminho)[1].lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(caminho, dtype=str)
    elif ext == ".csv":
        df = pd.read_csv(caminho, dtype=str, sep=None, engine="python")
    else:
        raise ValueError(f"Formato nao suportado: {ext}")

    renomear = {v: k for k, v in mapeamento.items() if v in df.columns}
    df = df.rename(columns=renomear)

    faltando = [k for k in ["numero_nf", "sku", "quantidade", "data"] if k not in df.columns]
    if faltando:
        raise ValueError(
            f"Colunas nao encontradas: {faltando}\n"
            f"Colunas no arquivo: {list(df.columns)}\n"
            f"Ajuste o mapeamento em config.py"
        )

    df["quantidade"] = pd.to_numeric(df["quantidade"].str.replace(",", "."), errors="coerce").fillna(0)
    if "peso" in df.columns:
        df["peso"] = pd.to_numeric(df["peso"].str.replace(",", "."), errors="coerce").fillna(0)
    if "vlr_unit" in df.columns:
        df["vlr_unit"] = pd.to_numeric(df["vlr_unit"].str.replace(",", "."), errors="coerce").fillna(0)
    # Detecta formato ISO (YYYY-MM-DD) vs brasileiro (DD/MM/YYYY) para não inverter mês/dia
    _sample = df["data"].dropna().head(10).tolist()
    _is_iso = any(re.match(r'^\d{4}[-/]\d{2}[-/]\d{2}', str(v)) for v in _sample)
    df["data"] = pd.to_datetime(df["data"], dayfirst=not _is_iso, errors="coerce")
    df["numero_nf"] = df["numero_nf"].str.strip()
    df["sku"]       = df["sku"].str.strip()

    if "status" in df.columns and WMS_FILTRO_STATUS:
        df = df[df["status"].isin(WMS_FILTRO_STATUS)]

    return df


# ── Auto-detecção de pares ────────────────────────────────────────────────────

def detectar_pares(pasta: str) -> list:
    """Retorna lista de (grupo, data_str, wms_path, erp_path) detectados na pasta."""
    pares = []
    for f in sorted(os.listdir(pasta)):
        m = re.match(r'^WMS (.+?)\s+(\d{1,2}[.\-]\d{2})\.xlsx$', f, re.IGNORECASE)
        if m:
            grupo    = m.group(1).strip()
            data_str = m.group(2).replace('-', '.')
        else:
            m2 = re.match(r'^WMS (.+?)\.xlsx$', f, re.IGNORECASE)
            if not m2:
                continue
            grupo    = m2.group(1).strip()
            data_str = ""
        sufixo   = f[4:]
        erp_path = os.path.join(pasta, "ERP " + sufixo)
        if os.path.exists(erp_path):
            pares.append((grupo, data_str, os.path.join(pasta, f), erp_path))
    return pares


# ── Conciliação ───────────────────────────────────────────────────────────────

def _cols_disponiveis(df: pd.DataFrame, desejadas: list) -> list:
    return [c for c in desejadas if c in df.columns]


def conciliar(df_wms: pd.DataFrame, df_erp: pd.DataFrame) -> dict:
    cols_wms = _cols_disponiveis(df_wms, ["numero_nf", "sku", "quantidade", "peso", "data", "descricao", "status"])
    cols_erp = _cols_disponiveis(df_erp, ["numero_nf", "sku", "quantidade", "unidade", "data", "descricao", "fornecedor", "vlr_unit"])

    df_wms = df_wms.drop_duplicates(subset=CHAVE_CONCILIACAO, keep="first")
    df_erp = df_erp.drop_duplicates(subset=CHAVE_CONCILIACAO, keep="first")

    merged = pd.merge(
        df_wms[cols_wms], df_erp[cols_erp],
        on=CHAVE_CONCILIACAO, how="outer",
        suffixes=("_wms", "_erp"), indicator=True,
    )

    so_wms = merged[merged["_merge"] == "left_only"].copy()
    so_erp = merged[merged["_merge"] == "right_only"].copy()
    ambos  = merged[merged["_merge"] == "both"].copy()

    is_kg = (
        ambos["unidade"].str.upper().str.strip() == "KG"
        if "unidade" in ambos.columns
        else pd.Series(False, index=ambos.index)
    )
    qtd_wms_ef = ambos["quantidade_wms"].copy()
    peso_col   = "peso_wms" if "peso_wms" in ambos.columns else ("peso" if "peso" in ambos.columns else None)
    if peso_col:
        qtd_wms_ef = qtd_wms_ef.where(~is_kg, ambos[peso_col])

    ambos["qtd_wms_usada"] = qtd_wms_ef
    ambos["diff_qtd"]      = (qtd_wms_ef - ambos["quantidade_erp"]).round(3)
    if "vlr_unit" in ambos.columns:
        ambos["valor_div"] = (ambos["diff_qtd"] * ambos["vlr_unit"]).round(2)

    ok      = ambos[ambos["diff_qtd"] == 0].copy()
    div_qtd = ambos[ambos["diff_qtd"] != 0].copy()

    return {
        "ok":      ok,
        "so_wms":  so_wms,
        "so_erp":  so_erp,
        "div_qtd": div_qtd,
    }


# ── Helpers HTML ──────────────────────────────────────────────────────────────

def _br_fmt(x, dec: int = 3, strip_zeros: bool = True) -> str:
    """Formata float em notação brasileira: 1.234,56"""
    if not pd.notna(x):
        return ""
    s = f"{abs(x):,.{dec}f}"          # ex: "10,316.080"
    if strip_zeros:
        intg, frac = s.split(".")
        frac = frac.rstrip("0")
        s = intg if not frac else f"{intg}.{frac}"   # "10,316.08"
    # troca separadores: , → . e . → ,
    br = s.replace(",", "\x00").replace(".", ",").replace("\x00", ".")
    return ("-" + br) if x < 0 else br


def _slug(texto: str) -> str:
    return re.sub(r"[^a-z0-9]", "_", texto.lower())


ROTULOS_HTML = {
    "numero_nf": "NF", "sku": "SKU",
    "descricao": "Descricao", "descricao_wms": "Descricao",
    "unidade": "Unidade", "status": "Status",
    "qtd_wms_usada": "Qtd WMS", "quantidade_wms": "Qtd WMS",
    "quantidade_erp": "Qtd ERP", "diff_qtd": "Diferenca",
    "peso": "Peso Liq.", "peso_wms": "Peso Liq.",
    "data_wms": "Data WMS", "data_erp": "Data ERP",
    "fornecedor": "Fornecedor",
}


def df_para_html(df: pd.DataFrame, cols: list) -> str:
    existentes = [c for c in cols if c in df.columns]
    if df.empty or not existentes:
        return '<p class="vazio">Nenhum registro nesta categoria.</p>'
    out = df[existentes].copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = out[col].dt.strftime("%d/%m/%Y").fillna("")
        elif pd.api.types.is_float_dtype(out[col]):
            out[col] = out[col].apply(lambda x: _br_fmt(x, 3) if pd.notna(x) else "")
    out.rename(columns=ROTULOS_HTML, inplace=True)
    return out.to_html(index=False, classes="tabela", border=0, na_rep="")


def _tabela_div_qtd_html(df: pd.DataFrame, cols: list, sg: str = "", data_str: str = "", grupo: str = "") -> str:
    """Tabela de divergência com botão de ajuste manual por linha."""
    existentes = [c for c in cols if c in df.columns]
    if df.empty or not existentes:
        return '<p class="vazio">Nenhuma divergencia encontrada.</p>'

    rotulos = {
        "numero_nf": "NF", "sku": "SKU",
        "descricao": "Descricao", "descricao_wms": "Descricao",
        "unidade": "Un.", "qtd_wms_usada": "Qtd WMS", "quantidade_erp": "Qtd ERP",
        "diff_qtd": "Diferenca", "valor_div": "Valor Div. (R$)",
        "data_wms": "Data", "fornecedor": "Fornecedor",
    }

    out = df[existentes].copy()

    # Salvar valores numéricos brutos antes de formatar (usados em data-attrs)
    raw_diff = out["diff_qtd"].to_dict()   if "diff_qtd"  in out.columns else {}
    raw_vdiv = out["valor_div"].to_dict()  if "valor_div" in out.columns else {}

    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = out[col].dt.strftime("%d/%m/%Y").fillna("")
        elif pd.api.types.is_float_dtype(out[col]):
            if col == "valor_div":
                out[col] = out[col].apply(lambda x: _br_fmt(x, 2, strip_zeros=False) if pd.notna(x) else "")
            else:
                out[col] = out[col].apply(lambda x: _br_fmt(x, 3) if pd.notna(x) else "")

    headers = "".join(f"<th>{rotulos.get(c, c)}</th>" for c in existentes) + "<th></th>"

    rows_html = ""
    for row_idx, row in out.iterrows():
        nf  = str(row.get("numero_nf", "")).strip()
        sku = str(row.get("sku", "")).strip()
        key = re.sub(r"[^a-zA-Z0-9]", "_", f"{nf}__{sku}")

        qtd_wms = str(row.get("qtd_wms_usada", ""))
        qtd_erp = str(row.get("quantidade_erp", ""))
        desc    = str(row.get("descricao", "")).replace("'", "").replace('"', "")[:60]

        # raw float do diff (antes da formatação BR) — para modal e data-attr
        rd         = raw_diff.get(row_idx, 0.0)
        rd         = 0.0 if (isinstance(rd, float) and rd != rd) else float(rd or 0)
        diff_str   = str(row.get("diff_qtd", ""))     # já em BR
        diff_modal = f"+{diff_str}" if rd > 0 else diff_str

        nf_js  = nf.replace("'", "\\'")
        sku_js = sku.replace("'", "\\'")

        cells = ""
        for c in existentes:
            v   = row[c]
            val = "" if (isinstance(v, float) and v != v) else ("" if v is None else str(v))
            if c == "diff_qtd":
                raw = raw_diff.get(row_idx, 0.0)
                raw = 0.0 if (isinstance(raw, float) and raw != raw) else float(raw or 0)
                display = f"+{val}" if raw > 0 else val
                css     = "td-diff td-diff-pos" if raw > 0 else "td-diff td-diff-neg"
                cells  += f'<td class="{css}" data-diff-original="{raw}">{display}</td>'
            elif c == "valor_div":
                raw = raw_vdiv.get(row_idx, 0.0)
                raw = 0.0 if (isinstance(raw, float) and raw != raw) else float(raw or 0)
                css   = "td-valordiv td-valordiv-pos" if raw > 0 else "td-valordiv td-valordiv-neg"
                cells += f'<td class="{css}">{val}</td>'
            else:
                cells += f"<td>{val}</td>"

        btn_cell = (
            f'<button class="btn-ajustar" '
            f'onclick="abrirAjuste(\'{key}\',\'{nf_js}\',\'{sku_js}\',\'{qtd_wms}\',\'{qtd_erp}\',\'{diff_modal}\',\'{desc}\')">'
            f'&#9998; Ajustar</button>'
            f'<span class="ajuste-badge" id="aj-{key}"></span>'
        )
        rows_html += (
            f'<tr data-ajuste-key="{key}" data-sg="{sg}" '
            f'data-data="{data_str}" data-grupo="{grupo}">'
            f'{cells}<td class="td-ajuste">{btn_cell}</td></tr>\n'
        )

    return f'<table class="tabela" border="0"><thead><tr>{headers}</tr></thead><tbody>{rows_html}</tbody></table>'


# ── Dashboard combinado (multi-grupo) ─────────────────────────────────────────

def _secao_grupo(grupo: str, resultados: dict, data_slug: str = "", data_str: str = "") -> str:
    ok      = resultados["ok"]
    so_wms  = resultados["so_wms"]
    so_erp  = resultados["so_erp"]
    div_qtd = resultados["div_qtd"]
    total_erp = len(ok) + len(so_erp) + len(div_qtd)
    pct_ok    = round(len(ok) / total_erp * 100, 1) if total_erp else 0

    def pct(n): return round(n / total_erp * 100, 1) if total_erp else 0

    sg = (data_slug + "_" if data_slug else "") + _slug(grupo)

    cols_ok      = ["numero_nf", "sku", "descricao", "unidade", "qtd_wms_usada", "quantidade_erp", "data_wms", "fornecedor"]
    cols_so_erp  = ["numero_nf", "sku", "descricao", "unidade", "quantidade_erp", "data_erp", "fornecedor"]
    cols_div_qtd = ["numero_nf", "sku", "descricao", "unidade", "qtd_wms_usada", "quantidade_erp", "diff_qtd", "valor_div", "data_wms", "fornecedor"]

    tab_ok      = df_para_html(ok,      cols_ok)
    tab_so_erp  = df_para_html(so_erp,  cols_so_erp)
    tab_div_qtd = _tabela_div_qtd_html(div_qtd, cols_div_qtd, sg, data_str, grupo)

    return f"""
<div id="grupo-{sg}" class="grupo-section" style="display:none;">

  <!-- Painel do grupo -->
  <div class="painel">
    <div class="donut-wrap">
      <div class="donut-container">
        <canvas id="donut-{sg}" width="180" height="180"></canvas>
        <div class="donut-centro">
          <div class="pct">{pct_ok}%</div>
          <div class="sub">Conciliado</div>
        </div>
      </div>
      <div class="donut-titulo">{grupo}</div>
    </div>
    <div class="painel-direita">
      <div class="painel-titulo">Resumo &nbsp;<span style="color:#555;font-size:12px;font-weight:400;">Base ERP: {total_erp} registros</span></div>
      <div class="metricas">
        <div class="metrica ok">
          <div class="m-numero">{len(ok)}</div>
          <div class="m-label">Conciliados OK</div>
          <div class="m-pct">{pct(len(ok))}%</div>
        </div>
        <div class="metrica so-erp">
          <div class="m-numero">{len(so_erp)}</div>
          <div class="m-label">So no ERP</div>
          <div class="m-pct">{pct(len(so_erp))}%</div>
        </div>
        <div class="metrica div-qtd">
          <div class="m-numero">{len(div_qtd)}</div>
          <div class="m-label">Div. Quantidade</div>
          <div class="m-pct">{pct(len(div_qtd))}%</div>
        </div>
      </div>
      <div class="barra-geral-wrap">
        <div class="barra-geral-label"><span>Distribuicao (base ERP)</span><span>{total_erp} registros</span></div>
        <div class="barra-geral">
          <div class="barra-seg ok"      style="width:{pct(len(ok))}%"></div>
          <div class="barra-seg so-erp"  style="width:{pct(len(so_erp))}%"></div>
          <div class="barra-seg div-qtd" style="width:{pct(len(div_qtd))}%"></div>
        </div>
        <div class="legenda">
          <div class="legenda-item"><div class="legenda-dot" style="background:#22c55e"></div> OK</div>
          <div class="legenda-item"><div class="legenda-dot" style="background:#3b82f6"></div> So ERP</div>
          <div class="legenda-item"><div class="legenda-dot" style="background:#ef4444"></div> Div Qtd</div>
        </div>
      </div>
    </div>
  </div>

  <!-- Abas do grupo -->
  <div class="tabs">
    <button class="tab-btn ativo" onclick="abrirAba('{sg}','ok',this)">
      Conciliados <span class="badge ok">{len(ok)}</span>
    </button>
    <button class="tab-btn" onclick="abrirAba('{sg}','so-erp',this)">
      So no ERP <span class="badge so-erp">{len(so_erp)}</span>
    </button>
    <button class="tab-btn" onclick="abrirAba('{sg}','div-qtd',this)">
      Div. Quantidade <span class="badge div-qtd">{len(div_qtd)}</span>
    </button>
  </div>

  <div id="{sg}-ok" class="tab-content ativo">
    <div class="acoes-wrap">
      <div class="filtro-wrap"><input type="text" placeholder="Filtrar..." onkeyup="filtrar(this,'{sg}-ok')"></div>
      <button class="btn-exportar" onclick="exportarCSV('{sg}-ok','{grupo}_conciliados_ok')">&#8595; Exportar CSV</button>
    </div>
    <div class="table-wrap">{tab_ok}</div>
  </div>
  <div id="{sg}-so-erp" class="tab-content" data-data="{data_str}" data-grupo="{grupo}">
    <div class="acoes-wrap">
      <div class="filtro-wrap"><input type="text" placeholder="Filtrar tabela..." onkeyup="filtrar(this,'{sg}-so-erp')"></div>
      <button class="btn-exportar" onclick="exportarCSV('{sg}-so-erp','{grupo}_so_no_erp')">&#8595; Exportar CSV</button>
    </div>
    <div class="busca-nf-wrap">
      <div class="busca-nf-header">
        <span class="busca-nf-icon">&#128270;</span>
        <input type="text" id="{sg}-busca-nf" class="busca-nf-input"
               placeholder="Buscar NF em todas as datas..."
               oninput="buscarNFOutrasDatas('{sg}', this.value)">
        <span class="busca-nf-hint">Digite o numero da NF para ver se foi recebida em outro dia</span>
      </div>
      <div id="{sg}-resultado-nf" class="resultado-nf" style="display:none;"></div>
    </div>
    <div class="table-wrap">{tab_so_erp}</div>
  </div>
  <div id="{sg}-div-qtd" class="tab-content">
    <div class="div-resumo" id="{sg}-dr">
      <div class="dr-item"><div class="dr-label">Total</div><strong id="{sg}-dr-total">{len(div_qtd)}</strong></div>
      <div class="dr-sep"></div>
      <div class="dr-item dr-pend"><div class="dr-label">Pendentes</div><strong id="{sg}-dr-pend">{len(div_qtd)}</strong></div>
      <div class="dr-item dr-parc"><div class="dr-label">Parcial</div><strong id="{sg}-dr-parc">0</strong></div>
      <div class="dr-item dr-ajust"><div class="dr-label">Ajustados</div><strong id="{sg}-dr-ajust">0</strong></div>
    </div>
    <div class="acoes-wrap">
      <div class="filtro-wrap"><input type="text" placeholder="Filtrar..." onkeyup="filtrar(this,'{sg}-div-qtd')"></div>
      <button class="btn-exportar" onclick="exportarCSV('{sg}-div-qtd','{grupo}_divergencia_qtd')">&#8595; Exportar CSV</button>
    </div>
    <div class="table-wrap">{tab_div_qtd}</div>
  </div>

</div>"""


def _donut_js(grupo: str, resultados: dict, data_slug: str = "") -> str:
    ok      = len(resultados["ok"])
    so_erp  = len(resultados["so_erp"])
    div_qtd = len(resultados["div_qtd"])
    sg = (data_slug + "_" if data_slug else "") + _slug(grupo)
    return f"""
  desenharDonut('donut-{sg}', [{ok},{so_erp},{div_qtd}]);"""


def gerar_dashboard_multi(todos: dict, data_ref: str, wms_files: list, erp_files: list) -> str:
    # todos = {data_str: {grupo: resultados}}
    datas  = sorted(todos.keys())
    todos_grupos = [g for grupos_d in todos.values() for g in grupos_d]

    # ── Botões de data ────────────────────────────────────────────────────────
    botoes_data = ""
    for d in datas:
        d_slug = _slug(d) if d else "sem_data"
        d_label = d.replace('.', '/') if d else "–"
        ativo = ' ativo' if d == datas[0] else ''
        botoes_data += f'<button class="data-btn{ativo}" onclick="abrirData(\'{d_slug}\')">{d_label}</button>\n'

    # ── Seções por data ───────────────────────────────────────────────────────
    secoes_data = ""
    init_js     = ""
    donuts_js   = ""

    for d in datas:
        d_slug   = _slug(d) if d else "sem_data"
        grupos_d = todos[d]
        grupos   = list(grupos_d.keys())

        ok_g  = sum(len(r["ok"])      for r in grupos_d.values())
        erp_g = sum(len(r["so_erp"]) for r in grupos_d.values())
        div_g = sum(len(r["div_qtd"]) for r in grupos_d.values())
        tot_g = ok_g + erp_g + div_g
        pct_g = round(ok_g / tot_g * 100, 1) if tot_g else 0

        # Cards por grupo desta data
        cards = ""
        for g, r in grupos_d.items():
            sg  = _slug(g)
            t   = len(r["ok"]) + len(r["so_erp"]) + len(r["div_qtd"])
            pct = round(len(r["ok"]) / t * 100, 1) if t else 0
            cor = "#22c55e" if pct == 100 else "#f59e0b" if pct >= 80 else "#f87171"
            cards += f"""
        <div class="card-grupo" onclick="abrirGrupo('{d_slug}','{sg}')">
          <div class="cg-nome">{g}</div>
          <div class="cg-pct" style="color:{cor};">{pct}%</div>
          <div class="cg-sub">OK</div>
          <div class="cg-detalhe">
            <span style="color:#22c55e;">✓ {len(r['ok'])}</span>
            <span style="color:#60a5fa;">E {len(r['so_erp'])}</span>
            <span style="color:#f87171;">D {len(r['div_qtd'])}</span>
          </div>
        </div>"""

        # Botões de grupo desta data
        btns_grupo = ""
        for g in grupos:
            sg    = _slug(g)
            ativo = ' ativo' if g == grupos[0] else ''
            btns_grupo += f'<button class="grupo-btn{ativo}" onclick="abrirGrupo(\'{d_slug}\',\'{sg}\')">{g}</button>\n'

        # Seções de grupo desta data
        secoes_grupos = "".join(_secao_grupo(g, r, d_slug, d) for g, r in grupos_d.items())
        donuts_js    += "".join(_donut_js(g, r, d_slug)    for g, r in grupos_d.items())

        primeiro_sg = _slug(grupos[0])
        init_js += f"abrirGrupo('{d_slug}','{primeiro_sg}');\n"

        display = '' if d == datas[0] else 'display:none;'
        secoes_data += f"""
<div id="data-section-{d_slug}" class="data-section" style="{display}">
  <div class="resumo-geral">
    <div class="rg-total">
      <strong>{ok_g}</strong> de <strong>{tot_g}</strong> itens ERP conciliados &mdash;
      <strong style="color:{'#22c55e' if pct_g==100 else '#f59e0b' if pct_g>=80 else '#f87171'};">{pct_g}%</strong> OK
    </div>
    <div style="flex:1;min-width:200px;">
      <div class="barra-geral">
        <div class="barra-seg ok"      style="width:{round(ok_g/tot_g*100,1) if tot_g else 0}%"></div>
        <div class="barra-seg so-erp"  style="width:{round(erp_g/tot_g*100,1) if tot_g else 0}%"></div>
        <div class="barra-seg div-qtd" style="width:{round(div_g/tot_g*100,1) if tot_g else 0}%"></div>
      </div>
    </div>
  </div>
  <div class="cards-grupos">{cards}</div>
  <div class="grupo-nav">{btns_grupo}</div>
  {secoes_grupos}
</div>"""

    todos_grupos_uniq = sorted(set(todos_grupos))

    # ── Dados para o painel de comparação ────────────────────────────────────
    grupos_uniq = sorted(set(g for grupos_d in todos.values() for g in grupos_d))
    dash_data   = []
    for d in datas:
        for g in grupos_uniq:
            r = todos[d].get(g)
            if not r:
                continue
            div_sum = float(r["div_qtd"]["diff_qtd"].sum()) if (not r["div_qtd"].empty and "diff_qtd" in r["div_qtd"].columns) else 0.0
            ok_n = len(r["ok"]); erp_n = len(r["so_erp"]); div_n = len(r["div_qtd"])
            tot  = ok_n + erp_n + div_n
            dash_data.append({
                "data": d, "grupo": g,
                "ok": ok_n, "erp": erp_n, "div": div_n, "total": tot,
                "div_sum": round(div_sum, 2),
                "pct_ok": round(ok_n / tot * 100, 1) if tot else 0,
            })
    dash_json   = json.dumps(dash_data,   ensure_ascii=False)
    grupos_json = json.dumps(grupos_uniq)
    datas_json  = json.dumps([d for d in datas if d])

    # ── Lookup cross-datas por NF (usado na busca da aba So no ERP) ───────────
    nf_lookup: dict = {}
    secao_map = [
        ("ok",      "ok",      "Conciliado"),
        ("so_wms",  "so-wms",  "So WMS"),
        ("so_erp",  "so-erp",  "So ERP"),
        ("div_qtd", "div-qtd", "Div. Qtd"),
    ]
    for d in datas:
        for g, r in todos[d].items():
            for secao_key, secao_css, secao_label in secao_map:
                df_s = r.get(secao_key, pd.DataFrame())
                if df_s.empty:
                    continue
                for _, row in df_s.iterrows():
                    nf = str(row.get("numero_nf", "")).strip()
                    if not nf:
                        continue
                    desc = str(row.get("descricao", ""))[:55]
                    sku  = str(row.get("sku", ""))
                    # qtd_wms: prefer qtd_wms_usada, fallback quantidade_wms
                    qw_raw = row.get("qtd_wms_usada", row.get("quantidade_wms"))
                    qe_raw = row.get("quantidade_erp")
                    qw = _br_fmt(float(qw_raw), 3) if pd.notna(qw_raw) and qw_raw != "" else ""
                    qe = _br_fmt(float(qe_raw), 3) if pd.notna(qe_raw) and qe_raw != "" else ""
                    entry = {
                        "data":    d,
                        "grupo":   g,
                        "secao":   secao_css,
                        "label":   secao_label,
                        "sku":     sku,
                        "desc":    desc,
                        "qtd_wms": qw,
                        "qtd_erp": qe,
                    }
                    nf_lookup.setdefault(nf, []).append(entry)
    nf_lookup_json = json.dumps(nf_lookup, ensure_ascii=False)

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Conciliacao de Recebimentos</title>
<style>
  * {{ box-sizing:border-box; margin:0; padding:0; }}
  body {{ font-family:'Segoe UI',sans-serif; background:#111111; color:#e0e0e0; }}

  header {{
    background:#1a1a1a; border-bottom:1px solid #2d2d2d;
    padding:18px 32px; display:flex; justify-content:space-between; align-items:center;
  }}
  header h1 {{ color:#f1f5f9; font-size:20px; font-weight:600; }}
  header .meta {{ font-size:12px; opacity:.55; text-align:right; line-height:1.7; }}
  .container {{ max-width:1400px; margin:0 auto; padding:24px 32px; }}

  /* Cards por grupo */
  .cards-grupos {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); gap:14px; margin-bottom:24px; }}
  .card-grupo {{
    background:#1e1e1e; border:1px solid #2d2d2d; border-radius:10px;
    padding:16px; text-align:center; cursor:pointer;
    transition:border-color .2s, background .2s;
  }}
  .card-grupo:hover {{ border-color:#6366f1; background:#252525; }}
  .card-grupo.ativo  {{ border-color:#6366f1; }}
  .cg-nome {{ font-size:12px; color:#888; margin-bottom:6px; font-weight:600; text-transform:uppercase; letter-spacing:.5px; }}
  .cg-pct  {{ font-size:30px; font-weight:700; line-height:1; }}
  .cg-sub  {{ font-size:10px; color:#555; margin-top:3px; }}
  .cg-detalhe {{ display:flex; justify-content:center; gap:8px; font-size:11px; font-weight:600; margin-top:8px; }}

  /* Resumo geral */
  .resumo-geral {{
    background:#1e1e1e; border:1px solid #2d2d2d; border-radius:12px;
    padding:16px 24px; margin-bottom:24px;
    display:flex; align-items:center; justify-content:space-between; flex-wrap:wrap; gap:16px;
  }}
  .rg-total {{ font-size:13px; color:#888; }}
  .rg-total strong {{ color:#f1f5f9; font-size:22px; }}
  .barra-geral {{ width:100%; max-width:600px; height:10px; background:#2a2a2a; border-radius:999px; overflow:hidden; display:flex; }}
  .barra-seg {{ height:100%; }}
  .barra-seg.ok      {{ background:#22c55e; }}
  .barra-seg.so-wms  {{ background:#f59e0b; }}
  .barra-seg.so-erp  {{ background:#60a5fa; }}
  .barra-seg.div-qtd {{ background:#f87171; }}

  /* Seletor de grupo */
  .grupo-nav {{ display:flex; gap:8px; flex-wrap:wrap; margin-bottom:20px; }}
  .grupo-btn {{
    padding:9px 22px; border:1px solid #2d2d2d; border-radius:8px;
    background:#1e1e1e; color:#888; font-size:13px; font-weight:600;
    cursor:pointer; transition:all .2s;
  }}
  .grupo-btn:hover {{ border-color:#6366f1; color:#818cf8; }}
  .grupo-btn.ativo  {{ background:#6366f1; border-color:#6366f1; color:#fff; }}

  /* Painel por grupo */
  .painel {{
    display:grid; grid-template-columns:240px 1fr; gap:24px;
    background:#1e1e1e; border:1px solid #2d2d2d; border-radius:14px;
    padding:24px; margin-bottom:20px; box-shadow:0 4px 20px rgba(0,0,0,.4);
  }}
  .donut-wrap {{ display:flex; flex-direction:column; align-items:center; justify-content:center; gap:10px; }}
  .donut-container {{ position:relative; width:180px; height:180px; }}
  .donut-container canvas {{ display:block; }}
  .donut-centro {{ position:absolute; top:50%; left:50%; transform:translate(-50%,-50%); text-align:center; }}
  .donut-centro .pct {{ font-size:30px; font-weight:700; color:#f1f5f9; line-height:1; }}
  .donut-centro .sub {{ font-size:11px; color:#666; margin-top:4px; }}
  .donut-titulo {{ font-size:12px; color:#666; text-align:center; }}
  .painel-direita {{ display:flex; flex-direction:column; justify-content:space-between; gap:14px; }}
  .painel-titulo {{ font-size:14px; font-weight:600; color:#ccc; border-bottom:1px solid #2d2d2d; padding-bottom:10px; }}
  .metricas {{ display:grid; grid-template-columns:repeat(4,1fr); gap:12px; }}
  .metrica {{ background:#191919; border-radius:8px; padding:14px; border-left:3px solid #333; }}
  .metrica .m-numero {{ font-size:26px; font-weight:700; line-height:1; }}
  .metrica .m-label  {{ font-size:11px; color:#666; margin-top:4px; }}
  .metrica .m-pct    {{ font-size:11px; font-weight:600; margin-top:5px; }}
  .metrica.ok      {{ border-color:#22c55e; }} .metrica.ok .m-numero,.metrica.ok .m-pct           {{ color:#22c55e; }}
  .metrica.so-wms  {{ border-color:#f59e0b; }} .metrica.so-wms .m-numero,.metrica.so-wms .m-pct   {{ color:#f59e0b; }}
  .metrica.so-erp  {{ border-color:#60a5fa; }} .metrica.so-erp .m-numero,.metrica.so-erp .m-pct   {{ color:#60a5fa; }}
  .metrica.div-qtd {{ border-color:#f87171; }} .metrica.div-qtd .m-numero,.metrica.div-qtd .m-pct {{ color:#f87171; }}
  .barra-geral-wrap {{ }}
  .barra-geral-label {{ display:flex; justify-content:space-between; font-size:11px; color:#555; margin-bottom:6px; }}
  .legenda {{ display:flex; gap:14px; flex-wrap:wrap; margin-top:8px; }}
  .legenda-item {{ display:flex; align-items:center; gap:6px; font-size:11px; color:#888; }}
  .legenda-dot {{ width:9px; height:9px; border-radius:50%; flex-shrink:0; }}

  /* Abas */
  .tabs {{ display:flex; gap:4px; flex-wrap:wrap; }}
  .tab-btn {{ padding:9px 18px; border:none; border-radius:8px 8px 0 0; cursor:pointer; font-size:13px; font-weight:600; background:#191919; color:#888; transition:background .2s,color .2s; }}
  .tab-btn:hover {{ background:#222; color:#ccc; }}
  .tab-btn.ativo {{ background:#1c1c1c; color:#818cf8; border-bottom:2px solid #6366f1; }}
  .tab-content {{ background:#1c1c1c; border-radius:0 8px 8px 8px; padding:20px; box-shadow:0 2px 12px rgba(0,0,0,.3); display:none; }}
  .tab-content.ativo {{ display:block; }}
  .acoes-wrap {{ display:flex; justify-content:space-between; align-items:center; margin-bottom:12px; gap:10px; flex-wrap:wrap; }}
  .filtro-wrap input {{ padding:7px 12px; border:1px solid #333; border-radius:6px; font-size:13px; outline:none; background:#141414; color:#e0e0e0; width:280px; }}
  .filtro-wrap input:focus {{ border-color:#6366f1; }}
  .filtro-wrap input::placeholder {{ color:#555; }}
  .btn-exportar {{ display:flex; align-items:center; gap:6px; padding:7px 16px; background:transparent; border:1px solid #6366f1; color:#818cf8; border-radius:6px; font-size:12px; font-weight:600; cursor:pointer; transition:background .2s,color .2s; white-space:nowrap; }}
  .btn-exportar:hover {{ background:#6366f1; color:#fff; }}
  .table-wrap {{ overflow-x:auto; }}
  .tabela {{ width:100%; border-collapse:collapse; font-size:13px; }}
  .tabela th {{ background:#1a1a1a; color:#94a3b8; padding:9px 12px; text-align:left; white-space:nowrap; border-bottom:1px solid #333; font-weight:600; letter-spacing:.3px; }}
  .tabela td {{ padding:8px 12px; border-bottom:1px solid #222; white-space:nowrap; color:#d0d0d0; }}
  .tabela tr:hover td {{ background:#222; }}
  .tabela tr:nth-child(even) td {{ background:#1a1a1a; }}
  .badge {{ display:inline-block; min-width:20px; padding:1px 5px; border-radius:10px; font-size:11px; font-weight:700; margin-left:5px; color:#fff; }}
  .badge.ok {{ background:#22c55e; }} .badge.so-wms {{ background:#f59e0b; }} .badge.so-erp {{ background:#60a5fa; }} .badge.div-qtd {{ background:#f87171; }}
  .vazio {{ color:#555; font-style:italic; padding:14px 0; }}
  footer {{ text-align:center; color:#444; font-size:11px; padding:28px 0; }}

  /* Seletor de data */
  .data-nav {{ display:flex; gap:8px; flex-wrap:wrap; margin-bottom:24px; padding-bottom:20px; border-bottom:1px solid #2d2d2d; }}
  .data-btn {{
    padding:10px 28px; border:1px solid #2d2d2d; border-radius:8px;
    background:#1a1a1a; color:#666; font-size:15px; font-weight:700;
    cursor:pointer; transition:all .2s; letter-spacing:.5px;
  }}
  .data-btn:hover {{ border-color:#6366f1; color:#818cf8; }}
  .data-btn.ativo  {{ background:#6366f1; border-color:#6366f1; color:#fff; }}

  /* Filtro por data de NF */
  .data-filtro {{ display:flex; align-items:center; gap:10px; margin-bottom:14px; flex-wrap:wrap; padding:12px 16px; background:#1a1a1a; border:1px solid #2d2d2d; border-radius:8px; }}
  .data-filtro span {{ font-size:12px; color:#666; }}
  .data-filtro input[type=date] {{ background:#141414; border:1px solid #333; border-radius:6px; color:#e0e0e0; font-size:12px; padding:6px 10px; outline:none; color-scheme:dark; }}
  .data-filtro input[type=date]:focus {{ border-color:#6366f1; }}
  .data-filtro button {{ background:transparent; border:1px solid #333; color:#666; border-radius:6px; font-size:11px; padding:6px 12px; cursor:pointer; transition:all .2s; }}
  .data-filtro button:hover {{ border-color:#6366f1; color:#818cf8; }}
  .btn-atualizar {{ display:flex; align-items:center; gap:6px; padding:8px 16px; background:transparent; border:1px solid #6366f1; color:#818cf8; border-radius:6px; font-size:12px; font-weight:600; cursor:pointer; transition:all .2s; }}
  .btn-atualizar:hover {{ background:#6366f1; color:#fff; }}

  /* Botão e badge de ajuste manual */
  .btn-ajustar {{ padding:4px 11px; background:transparent; border:1px solid #f59e0b; color:#f59e0b; border-radius:5px; font-size:11px; font-weight:600; cursor:pointer; transition:all .2s; white-space:nowrap; }}
  .btn-ajustar:hover {{ background:#f59e0b; color:#fff; }}
  .td-ajuste {{ white-space:nowrap; vertical-align:middle; }}
  .ajuste-badge {{ display:none; margin-left:8px; padding:2px 8px; border-radius:10px; font-size:10px; font-weight:700; background:#6366f1; color:#fff; vertical-align:middle; }}
  tr.tem-ajuste    > td:first-child {{ border-left:3px solid #f59e0b; }}
  tr.tem-parcial   > td:first-child {{ border-left:3px solid #f59e0b; }}
  tr.tem-ajustado  > td:first-child {{ border-left:3px solid #22c55e; }}
  tr.tem-ajustado  {{ opacity:.75; }}
  /* Divergência positiva (WMS > ERP) e negativa (WMS < ERP) */
  .td-diff-pos {{ color:#F59E0B; font-weight:700; }}
  .td-diff-neg {{ color:#F87171; font-weight:700; }}
  .td-valordiv {{ text-align:right; font-variant-numeric:tabular-nums; }}
  .td-valordiv-pos {{ color:#F59E0B; }}
  .td-valordiv-neg {{ color:#F87171; }}
  /* ── Busca NF cross-datas ── */
  .busca-nf-wrap {{ margin:0 0 14px 0; }}
  .busca-nf-header {{ display:flex; align-items:center; gap:10px; background:#0d1a2a; border:1px solid #1e3a5f; border-radius:8px; padding:10px 14px; }}
  .busca-nf-icon {{ font-size:16px; flex-shrink:0; }}
  .busca-nf-input {{ flex:1; background:transparent; border:none; outline:none; color:#e2e8f0; font-size:13px; min-width:0; }}
  .busca-nf-input::placeholder {{ color:#4a6080; }}
  .busca-nf-hint {{ font-size:11px; color:#4a6080; flex-shrink:0; white-space:nowrap; }}
  .resultado-nf {{ background:#0a1628; border:1px solid #1e3a5f; border-top:none; border-radius:0 0 8px 8px; overflow-x:auto; }}
  .resultado-nf .nf-vazio {{ padding:12px 16px; color:#4a6080; font-size:12px; font-style:italic; }}
  .resultado-nf table {{ width:100%; border-collapse:collapse; font-size:12px; }}
  .resultado-nf th {{ background:#0d1a2a; color:#60a5fa; font-weight:600; padding:7px 10px; text-align:left; border-bottom:1px solid #1e3a5f; white-space:nowrap; }}
  .resultado-nf td {{ padding:6px 10px; border-bottom:1px solid #0d1a2a; color:#cbd5e1; vertical-align:middle; }}
  .resultado-nf tr:last-child td {{ border-bottom:none; }}
  .resultado-nf tr:hover td {{ background:#0d1a2a; }}
  .badge-secao {{ display:inline-block; padding:2px 7px; border-radius:4px; font-size:11px; font-weight:700; }}
  .badge-secao.ok {{ background:#14532d; color:#22c55e; }}
  .badge-secao.so-wms {{ background:#172554; color:#60a5fa; }}
  .badge-secao.so-erp {{ background:#1e1b4b; color:#a78bfa; }}
  .badge-secao.div-qtd {{ background:#450a0a; color:#f87171; }}

  /* Modal */
  .modal-overlay {{ display:none; position:fixed; inset:0; background:rgba(0,0,0,.75); z-index:1000; align-items:center; justify-content:center; }}
  .modal-overlay.ativo {{ display:flex; }}
  .modal {{ background:#1e1e1e; border:1px solid #6366f1; border-radius:14px; padding:28px 32px; min-width:400px; max-width:520px; width:90%; box-shadow:0 8px 40px rgba(99,102,241,.2); }}
  .modal-titulo {{ color:#818cf8; font-size:16px; font-weight:700; margin-bottom:4px; }}
  .modal-info {{ font-size:12px; color:#555; margin-bottom:6px; }}
  .modal-qtds {{ display:flex; gap:20px; font-size:12px; margin-bottom:20px; padding:10px 14px; background:#141414; border-radius:8px; border:1px solid #2d2d2d; }}
  .modal label {{ display:block; font-size:12px; color:#888; margin-bottom:5px; margin-top:14px; font-weight:600; }}
  .modal input, .modal select, .modal textarea {{ width:100%; background:#141414; border:1px solid #333; border-radius:6px; color:#e0e0e0; font-size:13px; padding:9px 12px; outline:none; box-sizing:border-box; color-scheme:dark; font-family:inherit; }}
  .modal input:focus, .modal select:focus, .modal textarea:focus {{ border-color:#6366f1; }}
  .modal select option {{ background:#141414; }}
  .modal textarea {{ resize:vertical; min-height:65px; }}
  .modal-btns {{ display:flex; gap:10px; margin-top:22px; justify-content:flex-end; }}
  .modal-btns button {{ padding:9px 24px; border-radius:6px; font-size:13px; font-weight:600; cursor:pointer; transition:all .2s; }}
  .btn-salvar {{ background:#6366f1; color:#fff; border:none; }}
  .btn-salvar:hover {{ background:#4f46e5; }}
  .btn-cancelar {{ background:transparent; color:#888; border:1px solid #333; }}
  .btn-cancelar:hover {{ border-color:#6366f1; color:#818cf8; }}
  /* Toggle desconsiderar */
  .toggle-desc {{ display:flex; align-items:center; gap:10px; margin:16px 0 4px 0; padding:10px 14px; background:#1a0a0a; border:1px solid #ef4444; border-radius:8px; cursor:pointer; user-select:none; }}
  .toggle-desc input {{ display:none; }}
  .tg-track {{ width:38px; height:20px; background:#333; border-radius:10px; position:relative; transition:background .25s; flex-shrink:0; }}
  .tg-thumb {{ position:absolute; width:16px; height:16px; background:#666; border-radius:50%; top:2px; left:2px; transition:all .25s; }}
  .toggle-desc input:checked + .tg-track {{ background:#ef4444; }}
  .toggle-desc input:checked + .tg-track .tg-thumb {{ left:20px; background:#fff; }}
  .toggle-desc .tg-label {{ font-size:12px; color:#f87171; font-weight:700; line-height:1.3; }}
  .modal input:disabled {{ opacity:.45; cursor:not-allowed; }}

  /* Resumo divergências */
  .div-resumo {{ display:flex; gap:0; margin-bottom:14px; background:#141414; border:1px solid #2d2d2d; border-radius:10px; overflow:hidden; }}
  .dr-item {{ flex:1; display:flex; flex-direction:column; align-items:center; padding:12px 8px; }}
  .dr-sep {{ width:1px; background:#2d2d2d; margin:8px 0; }}
  .dr-label {{ font-size:10px; color:#666; text-transform:uppercase; letter-spacing:.6px; margin-bottom:4px; }}
  .dr-item strong {{ font-size:22px; font-weight:700; color:#e0e0e0; }}
  .dr-pend strong {{ color:#f87171; }}
  .dr-parc strong {{ color:#f59e0b; }}
  .dr-ajust strong {{ color:#22c55e; }}

  /* Toggle análise / dashboard */
  .view-toggle {{ display:flex; gap:4px; background:#141414; padding:3px; border-radius:8px; border:1px solid #2d2d2d; }}
  .view-btn {{ padding:6px 16px; border:none; border-radius:6px; font-size:12px; font-weight:600; cursor:pointer; color:#666; background:transparent; transition:all .2s; }}
  .view-btn.ativo {{ background:#6366f1; color:#fff; }}

  /* Export consolidado */
  .btn-export-cons {{ display:flex; align-items:center; gap:6px; padding:8px 16px; background:transparent; border:1px solid #22c55e; color:#22c55e; border-radius:6px; font-size:12px; font-weight:600; cursor:pointer; transition:all .2s; white-space:nowrap; }}
  .btn-export-cons:hover {{ background:#22c55e; color:#fff; }}

  /* Painel dashboard */
  .dash-panel {{ display:none; }}
  .dash-panel.ativo {{ display:block; }}
  .dash-cards {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(200px,1fr)); gap:14px; margin-bottom:20px; }}
  .dash-card {{ background:#1e1e1e; border:1px solid #2d2d2d; border-radius:12px; padding:20px 24px; }}
  .dash-card-label {{ font-size:10px; color:#666; text-transform:uppercase; letter-spacing:.6px; margin-bottom:8px; }}
  .dash-card-valor {{ font-size:32px; font-weight:700; color:#f1f5f9; line-height:1; }}
  .dash-card-sub {{ font-size:12px; color:#666; margin-top:6px; }}
  .dash-secao {{ background:#1e1e1e; border:1px solid #2d2d2d; border-radius:12px; padding:20px 24px; margin-bottom:16px; }}
  .dash-secao-titulo {{ font-size:13px; font-weight:600; color:#94a3b8; margin-bottom:16px; text-transform:uppercase; letter-spacing:.5px; font-size:11px; }}
  .dash-tabela {{ width:100%; border-collapse:collapse; font-size:13px; }}
  .dash-tabela th {{ background:#141414; color:#94a3b8; padding:9px 14px; text-align:left; border-bottom:1px solid #333; font-weight:600; white-space:nowrap; }}
  .dash-tabela td {{ padding:9px 14px; border-bottom:1px solid #222; color:#d0d0d0; white-space:nowrap; }}
  .dash-tabela tr:hover td {{ background:#222; }}
  .dash-tabela .pct-bar {{ display:inline-block; height:6px; border-radius:3px; vertical-align:middle; margin-left:8px; }}
  .analise-panel {{ display:none; }}
  .analise-panel.ativo {{ display:block; }}
</style>
</head>
<body>

<header>
  <div style="display:flex;align-items:center;gap:16px;">
    <a href="index.html" style="display:flex;align-items:center;gap:6px;color:#818cf8;text-decoration:none;font-size:13px;font-weight:600;border:1px solid #6366f1;border-radius:6px;padding:6px 12px;"
       onmouseover="this.style.background='#6366f1';this.style.color='#fff'"
       onmouseout="this.style.background='transparent';this.style.color='#818cf8'">&#8592; Inicio</a>
    <div>
      <h1>Conciliacao de Recebimentos</h1>
      <div style="font-size:13px;opacity:.7;margin-top:3px;color:#aaa;">WMS vs ERP &mdash; {data_ref}</div>
    </div>
  </div>
  <div style="display:flex;align-items:center;gap:16px;">
    <div class="meta" style="text-align:right;">
      <div>Grupos: <strong style="color:#ddd;">{" | ".join(todos_grupos_uniq)}</strong></div>
      <div>Gerado em {datetime.now().strftime("%d/%m/%Y %H:%M")}</div>
    </div>
    <div class="view-toggle">
      <button class="view-btn ativo" id="vbtn-analise"  onclick="alternarView('analise')">&#128203; Analise</button>
      <button class="view-btn"       id="vbtn-dashboard" onclick="alternarView('dashboard')">&#128202; Dashboard</button>
    </div>
    <button class="btn-export-cons" id="btn-exportar-div" onclick="exportarDivergenciasConsolidadas()">&#11015; Exportar Divergencias</button>
    <button class="btn-export-cons" id="btn-exportar-soerp" onclick="exportarSoERPConsolidado()" style="background:#1e3a5f;border-color:#3b82f6;color:#93c5fd;">&#11015; Exportar So ERP</button>
    <button class="btn-atualizar" id="btn-atualizar" onclick="atualizarDados()">&#8635; Atualizar</button>
  </div>
</header>

<div class="container">

<!-- ── PAINEL ANALISE ──────────────────────────────────────────────────── -->
<div id="painel-analise" class="analise-panel ativo">

  <!-- Seletor de data -->
  {'<div class="data-nav">' + botoes_data + '</div>' if len(datas) > 1 else ''}

  <!-- Seções por data -->
  {secoes_data}

</div>

<!-- ── PAINEL DASHBOARD ────────────────────────────────────────────────── -->
<div id="painel-dashboard" class="dash-panel">
  <div class="dash-cards" id="dash-cards-sumario"></div>
  <div class="dash-secao">
    <div class="dash-secao-titulo">% Conciliacao OK por Cliente</div>
    <canvas id="dash-bar-canvas" style="width:100%;max-height:320px;"></canvas>
  </div>
  <div class="dash-secao">
    <div class="dash-secao-titulo">Detalhe por Grupo</div>
    <table class="dash-tabela" id="dash-tabela-detalhe">
      <thead><tr>
        <th>Data</th><th>Grupo</th><th>OK</th><th>So ERP</th><th>Divergencias</th><th>Dif. Qtd.</th><th>% OK</th>
      </tr></thead>
      <tbody id="dash-tbody"></tbody>
    </table>
  </div>
</div>

</div>

<footer>Conciliacao de Estoque &mdash; {datetime.now().strftime("%d/%m/%Y %H:%M")}</footer>

<!-- Modal de ajuste manual -->
<div class="modal-overlay" id="modal-ajuste" onclick="if(event.target===this)fecharAjuste()">
  <div class="modal">
    <div class="modal-titulo">&#9998; Ajuste Manual</div>
    <div class="modal-info">NF: <strong id="modal-nf" style="color:#ddd"></strong> &nbsp;|&nbsp; SKU: <strong id="modal-sku" style="color:#ddd"></strong></div>
    <div class="modal-qtds" id="modal-qtds"></div>
    <label class="toggle-desc" onclick="toggleDesconsiderar()">
      <input type="checkbox" id="aj-desconsiderar">
      <span class="tg-track"><span class="tg-thumb"></span></span>
      <span class="tg-label">Desconsiderar divergencia &mdash; ajustar falta/sobra total automaticamente</span>
    </label>
    <label>Quantidade ajuste</label>
    <input type="number" id="aj-quantidade" step="any" placeholder="Ex: 10">
    <label>Motivo</label>
    <select id="aj-motivo">
      <option value="">Selecione...</option>
      <option value="Desconsiderado">Diferenca desconsiderada</option>
      <option value="Devolucao">Devolucao ao fornecedor</option>
      <option value="Erro WMS">Erro lancamento WMS</option>
      <option value="Erro ERP">Erro lancamento ERP</option>
      <option value="Avaria">Avaria / Perda</option>
      <option value="Reprocessamento">Reprocessamento</option>
      <option value="Outros">Outros</option>
    </select>
    <label>Data de emissao</label>
    <input type="date" id="aj-data-emissao">
    <label>Observacao (opcional)</label>
    <textarea id="aj-obs" placeholder="Detalhes adicionais..."></textarea>
    <div class="modal-btns">
      <button class="btn-cancelar" onclick="fecharAjuste()">Cancelar</button>
      <button class="btn-salvar" onclick="salvarAjuste()">&#10003; Salvar</button>
    </div>
  </div>
</div>

<script>
// ── Helpers de formato numérico BR ───────────────────────────────────────────
function parseBR(s) {{
  if (!s) return 0;
  const t = String(s).trim();
  // período = separador de milhar, vírgula = decimal → remove pontos, troca vírgula por ponto
  return parseFloat(t.replace(/\./g,'').replace(',','.')) || 0;
}}
function toBR(num, dec) {{
  if (isNaN(num) || num === null) return '0';
  const d = dec !== undefined ? dec : 3;
  const fixed = Math.abs(num).toFixed(d);
  const [intP, decP] = fixed.split('.');
  const intFmt = intP.replace(/\B(?=(\d{{3}})+(?!\d))/g, '.');
  const decStripped = decP ? decP.replace(/0+$/,'') : '';
  const result = decStripped ? intFmt + ',' + decStripped : intFmt;
  return num < 0 ? '-' + result : result;
}}

function desenharDonut(id, vals) {{
  const cores = ['#22c55e','#60a5fa','#f87171'];
  const canvas = document.getElementById(id);
  if (!canvas) return;
  const ctx = canvas.getContext('2d');
  const cx=90,cy=90,r=72,inner=48;
  const total = vals.reduce((s,v)=>s+v,0);
  let ang = -Math.PI/2;
  ctx.clearRect(0,0,180,180);
  if (total===0) {{
    ctx.beginPath(); ctx.arc(cx,cy,r,0,Math.PI*2); ctx.fillStyle='#2a2a2a'; ctx.fill();
    ctx.beginPath(); ctx.arc(cx,cy,inner,0,Math.PI*2); ctx.fillStyle='#1e1e1e'; ctx.fill();
    return;
  }}
  vals.forEach((v,i) => {{
    if(v===0) return;
    const fatia=(v/total)*Math.PI*2;
    ctx.beginPath(); ctx.moveTo(cx,cy); ctx.arc(cx,cy,r,ang,ang+fatia); ctx.closePath();
    ctx.fillStyle=cores[i]; ctx.fill(); ang+=fatia;
  }});
  ctx.beginPath(); ctx.arc(cx,cy,inner,0,Math.PI*2); ctx.fillStyle='#1e1e1e'; ctx.fill();
}}

function abrirData(d) {{
  document.querySelectorAll('.data-section').forEach(el=>el.style.display='none');
  document.querySelectorAll('.data-btn').forEach(el=>el.classList.remove('ativo'));
  document.getElementById('data-section-'+d).style.display='block';
  document.querySelectorAll('.data-btn').forEach(btn=>{{
    if(btn.getAttribute('onclick').includes("'"+d+"'")) btn.classList.add('ativo');
  }});
}}

function abrirGrupo(d, sg) {{
  const fullSg = d + '_' + sg;
  const sec = document.getElementById('data-section-'+d);
  sec.querySelectorAll('.grupo-section').forEach(el=>el.style.display='none');
  sec.querySelectorAll('.grupo-btn').forEach(el=>el.classList.remove('ativo'));
  sec.querySelectorAll('.card-grupo').forEach(el=>el.classList.remove('ativo'));
  document.getElementById('grupo-'+fullSg).style.display='block';
  sec.querySelectorAll('.grupo-btn').forEach(btn=>{{
    if(btn.getAttribute('onclick').includes("'"+sg+"'")) btn.classList.add('ativo');
  }});
  sec.querySelectorAll('.card-grupo').forEach(c=>{{
    if(c.getAttribute('onclick') && c.getAttribute('onclick').includes("'"+sg+"'")) c.classList.add('ativo');
  }});
}}

function abrirAba(sg, id, btn) {{
  document.querySelectorAll('#grupo-'+sg+' .tab-content').forEach(el=>el.classList.remove('ativo'));
  document.querySelectorAll('#grupo-'+sg+' .tab-btn').forEach(el=>el.classList.remove('ativo'));
  document.getElementById(sg+'-'+id).classList.add('ativo');
  btn.classList.add('ativo');
}}

function filtrar(input, abaId) {{
  const t=input.value.toLowerCase();
  document.querySelectorAll('#'+abaId+' .tabela tbody tr').forEach(tr=>{{
    tr.style.display=tr.textContent.toLowerCase().includes(t)?'':'none';
  }});
}}

// ── Busca NF em todas as datas (aba So no ERP) ───────────────────────────────
function buscarNFOutrasDatas(sg, val) {{
  const el = document.getElementById(sg + '-resultado-nf');
  if (!el) return;
  const q = val.trim().replace(/^0+/, '');   // remove zeros à esquerda para busca flexível
  if (q.length < 3) {{
    el.style.display = 'none';
    el.innerHTML = '';
    return;
  }}

  const encontrados = [];
  for (const [nf, ocorrencias] of Object.entries(NF_LOOKUP)) {{
    const nfNum = nf.replace(/^0+/, '');
    if (nfNum.includes(q) || nf.includes(val.trim())) {{
      ocorrencias.forEach(o => encontrados.push({{nf, ...o}}));
    }}
  }}

  if (encontrados.length === 0) {{
    el.innerHTML = '<div class="nf-vazio">Nenhuma ocorrencia encontrada para esta NF em nenhuma data ou secao.</div>';
    el.style.display = 'block';
    return;
  }}

  // Ordena: primeiro as de outras seções (não so-erp), depois por data
  encontrados.sort((a, b) => {{
    const p = (s) => s === 'so-erp' ? 1 : 0;
    return p(a.secao) - p(b.secao) || a.data.localeCompare(b.data);
  }});

  let html = '<table><thead><tr>'
    + '<th>NF</th><th>Data</th><th>Grupo</th><th>Secao</th>'
    + '<th>SKU</th><th>Descricao</th><th>Qtd WMS</th><th>Qtd ERP</th>'
    + '</tr></thead><tbody>';
  encontrados.forEach(r => {{
    const dataFmt = r.data ? r.data.replace('.', '/') : '-';
    html += `<tr>
      <td>${{r.nf}}</td>
      <td>${{dataFmt}}</td>
      <td>${{r.grupo}}</td>
      <td><span class="badge-secao ${{r.secao}}">${{r.label}}</span></td>
      <td>${{r.sku}}</td>
      <td style="max-width:220px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;">${{r.desc}}</td>
      <td style="text-align:right;">${{r.qtd_wms || '—'}}</td>
      <td style="text-align:right;">${{r.qtd_erp || '—'}}</td>
    </tr>`;
  }});
  html += '</tbody></table>';
  el.innerHTML = html;
  el.style.display = 'block';
}}

function filtrarPorData(sg) {{
  const de  = document.getElementById(sg+'-data-de').value;
  const ate = document.getElementById(sg+'-data-ate').value;
  const deDt  = de  ? new Date(de+'T00:00:00')  : null;
  const ateDt = ate ? new Date(ate+'T23:59:59') : null;
  document.querySelectorAll('#grupo-'+sg+' .tabela tbody tr').forEach(tr => {{
    if (!deDt && !ateDt) {{ tr.style.display=''; return; }}
    let dtCelula = null;
    tr.querySelectorAll('td').forEach(td => {{
      if (/^\\d{{2}}\\/\\d{{2}}\\/\\d{{4}}$/.test(td.textContent.trim())) {{
        const [d,m,y] = td.textContent.trim().split('/');
        dtCelula = new Date(y+'-'+m+'-'+d+'T00:00:00');
      }}
    }});
    if (!dtCelula) {{ tr.style.display=''; return; }}
    const ok = (!deDt || dtCelula >= deDt) && (!ateDt || dtCelula <= ateDt);
    tr.style.display = ok ? '' : 'none';
  }});
}}

function limparFiltroData(sg) {{
  document.getElementById(sg+'-data-de').value  = '';
  document.getElementById(sg+'-data-ate').value = '';
  document.querySelectorAll('#grupo-'+sg+' .tabela tbody tr').forEach(tr => tr.style.display='');
}}

function exportarCSV(abaId, nome) {{
  const tabela=document.querySelector('#'+abaId+' .tabela');
  if(!tabela) return;
  const linhas=[...tabela.querySelectorAll('tr')].filter(tr=>tr.style.display!=='none');
  const csv=linhas.map(tr=>[...tr.querySelectorAll('th,td')].map(c=>'"'+c.textContent.replace(/"/g,'""').trim()+'"').join(';')).join('\\n');
  const a=document.createElement('a');
  a.href=URL.createObjectURL(new Blob(['\\uFEFF'+csv],{{type:'text/csv;charset=utf-8;'}}));
  a.download=nome+'.csv'; a.click(); URL.revokeObjectURL(a.href);
}}

// ── Ajuste manual ────────────────────────────────────────────────────────────
let _ajKey = '';
let _ajDiffAbs = 0;

function abrirAjuste(key, nf, sku, qtdWms, qtdErp, diff, desc) {{
  _ajKey = key;
  _ajDiffAbs = Math.abs(parseBR(diff) || 0);
  document.getElementById('modal-nf').textContent  = nf;
  document.getElementById('modal-sku').textContent = sku;
  document.getElementById('modal-qtds').innerHTML  =
    '<span>WMS: <strong style="color:#22c55e">' + qtdWms + '</strong></span>' +
    '<span>ERP: <strong style="color:#3b82f6">' + qtdErp + '</strong></span>' +
    '<span>Dif: <strong style="color:#ef4444">' + diff  + '</strong></span>' +
    (desc ? '<span style="color:#555;font-size:11px;flex:1;text-align:right;overflow:hidden;text-overflow:ellipsis;">' + desc + '</span>' : '');
  const saved = JSON.parse(localStorage.getItem('ajuste_' + key) || 'null');
  const isDesc = saved && saved.desconsiderado;
  document.getElementById('aj-desconsiderar').checked = !!isDesc;
  document.getElementById('aj-quantidade').value   = isDesc ? _ajDiffAbs : (saved ? (saved.quantidade || '') : '');
  document.getElementById('aj-quantidade').disabled = !!isDesc;
  document.getElementById('aj-motivo').value       = saved ? (saved.motivo    || '') : '';
  document.getElementById('aj-data-emissao').value = saved ? (saved.data      || '') : '';
  document.getElementById('aj-obs').value          = saved ? (saved.obs       || '') : '';
  document.getElementById('modal-ajuste').classList.add('ativo');
}}

function toggleDesconsiderar() {{
  const chk = document.getElementById('aj-desconsiderar');
  chk.checked = !chk.checked;
  const qtdEl    = document.getElementById('aj-quantidade');
  const motivoEl = document.getElementById('aj-motivo');
  if (chk.checked) {{
    qtdEl.value    = _ajDiffAbs;
    qtdEl.disabled = true;
    if (!motivoEl.value) motivoEl.value = 'Desconsiderado';
  }} else {{
    qtdEl.value    = '';
    qtdEl.disabled = false;
    if (motivoEl.value === 'Desconsiderado') motivoEl.value = '';
  }}
}}

function fecharAjuste() {{
  document.getElementById('modal-ajuste').classList.remove('ativo');
}}

function salvarAjuste() {{
  const motivo = document.getElementById('aj-motivo').value;
  if (!motivo) {{ alert('Selecione um motivo.'); return; }}
  const isDesc = document.getElementById('aj-desconsiderar').checked;
  const qtdEl  = document.getElementById('aj-quantidade');
  qtdEl.disabled = false;
  const dados = {{
    quantidade:     isDesc ? String(_ajDiffAbs) : qtdEl.value,
    motivo:         motivo,
    data:           document.getElementById('aj-data-emissao').value,
    obs:            document.getElementById('aj-obs').value,
    desconsiderado: isDesc,
  }};
  localStorage.setItem('ajuste_' + _ajKey, JSON.stringify(dados));
  _aplicarBadge(_ajKey, dados);
  fecharAjuste();
}}

function _aplicarBadge(key, dados) {{
  // Badge de motivo
  const badge = document.getElementById('aj-' + key);
  if (badge) {{ badge.textContent = dados.motivo; badge.style.display = 'inline-block'; }}

  // Deducao na celula diff_qtd
  const row = document.querySelector('tr[data-ajuste-key="' + key + '"]');
  if (!row) return;
  const tdDiff = row.querySelector('.td-diff');
  if (tdDiff) {{
    const original = parseFloat(tdDiff.getAttribute('data-diff-original') || '0');
    const origBR   = (original >= 0 ? '+' : '-') + toBR(Math.abs(original), 3);
    const ajuste   = parseFloat(dados.quantidade || '0');
    const origAbs  = Math.abs(original);
    const restante = origAbs - Math.abs(ajuste);
    if (restante <= 0) {{
      tdDiff.innerHTML = '<span style="text-decoration:line-through;color:#555;font-size:11px;">' + origBR + '</span> <span style="color:#22c55e;font-weight:700;">0</span>';
      row.classList.remove('tem-ajuste','tem-parcial');
      row.classList.add('tem-ajustado');
    }} else {{
      const sinal = original < 0 ? '-' : '+';
      tdDiff.innerHTML = '<span style="text-decoration:line-through;color:#555;font-size:11px;">' + origBR + '</span> <span style="color:#f59e0b;font-weight:700;">' + sinal + toBR(restante, 3) + '</span>';
      row.classList.remove('tem-ajustado');
      row.classList.add('tem-ajuste','tem-parcial');
    }}
  }}

  const sg = row.getAttribute('data-sg');
  if (sg) atualizarResumoDivergencia(sg);
}}

function atualizarResumoDivergencia(sg) {{
  const rows = document.querySelectorAll('tr[data-sg="' + sg + '"]');
  let pend=0, parc=0, ajust=0;
  rows.forEach(r => {{
    if      (r.classList.contains('tem-ajustado')) ajust++;
    else if (r.classList.contains('tem-parcial'))  parc++;
    else                                            pend++;
  }});
  const el = id => document.getElementById(sg + '-dr-' + id);
  if (el('pend'))  el('pend').textContent  = pend;
  if (el('parc'))  el('parc').textContent  = parc;
  if (el('ajust')) el('ajust').textContent = ajust;
}}

function carregarAjustes() {{
  for (let i = 0; i < localStorage.length; i++) {{
    const lsKey = localStorage.key(i);
    if (!lsKey || !lsKey.startsWith('ajuste_')) continue;
    const key   = lsKey.replace('ajuste_', '');
    const dados = JSON.parse(localStorage.getItem(lsKey) || 'null');
    if (dados) _aplicarBadge(key, dados);
  }}
}}

// ── Exportar divergencias consolidadas ───────────────────────────────────────
function exportarDivergenciasConsolidadas() {{
  const rows = document.querySelectorAll('tr[data-ajuste-key]');
  const linhas = [['Data','Grupo','NF','SKU','Descricao','Qtd WMS','Qtd ERP','Diff Original','Diff Restante','Valor Div. (R$)','Valor Ajustado (R$)','Status','Motivo Ajuste','Qtd Ajuste','Data Emissao','Observacao']];
  rows.forEach(row => {{
    if (row.style.display === 'none') return;
    const key   = row.getAttribute('data-ajuste-key');
    const data  = row.getAttribute('data-data')  || '';
    const grupo = row.getAttribute('data-grupo') || '';
    const tds   = row.querySelectorAll('td');
    const vals  = [...tds].map(td => td.textContent.trim());
    const tdDiff = row.querySelector('.td-diff');
    const original = tdDiff ? (tdDiff.getAttribute('data-diff-original') || '') : '';
    const saved = JSON.parse(localStorage.getItem('ajuste_' + key) || 'null');
    const motivo   = saved ? (saved.motivo    || '') : '';
    const qtdAj    = saved ? (saved.quantidade || '') : '';
    const dtEmissao= saved ? (saved.data       || '') : '';
    const obs      = saved ? (saved.obs        || '') : '';
    let status = 'Pendente';
    if (row.classList.contains('tem-ajustado')) status = 'Ajustado';
    else if (row.classList.contains('tem-parcial')) status = 'Parcial';
    const restEl = tdDiff ? tdDiff.querySelector('span:last-child') : null;
    const restante = restEl ? restEl.textContent.trim() : original;
    // cols: NF=0, SKU=1, Descricao=2, Unid=3, WMS=4, ERP=5, Diff=6, ValorDiv=7, Data=8, Forn=9
    const nf      = vals[0] || '';
    const sku     = vals[1] || '';
    const desc    = vals[2] || '';
    const qtdWms  = vals[4] || '';
    const qtdErp  = vals[5] || '';
    const valorDivNum = parseBR(vals[7] || '0');
    const origNum     = parseFloat(original || '0') || 0;
    const vlrUnit     = origNum !== 0 ? valorDivNum / origNum : 0;
    const valorDiv    = valorDivNum.toFixed(2);
    const valorAjust  = (parseFloat(qtdAj || '0') * Math.abs(vlrUnit)).toFixed(2);
    linhas.push([data, grupo, nf, sku, desc, qtdWms, qtdErp, original, restante, valorDiv, valorAjust, status, motivo, qtdAj, dtEmissao, obs]);
  }});
  const headers = linhas.shift();
  const btn = document.getElementById('btn-exportar-div');
  if (btn) {{ btn.disabled = true; btn.textContent = '⏳ Gerando...'; }}
  fetch('/api/exportar_divergencias', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/json'}},
    body: JSON.stringify({{headers: headers, rows: linhas}})
  }})
  .then(r => r.blob())
  .then(blob => {{
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url;
    a.download = 'divergencias_consolidadas.xlsx';
    a.click();
    URL.revokeObjectURL(url);
  }})
  .catch(e => alert('Erro ao gerar Excel: ' + e))
  .finally(() => {{
    if (btn) {{ btn.disabled = false; btn.textContent = '📥 Exportar Divergências'; }}
  }});
}}

// ── Exportar So ERP consolidado ──────────────────────────────────────────────
function exportarSoERPConsolidado() {{
  // cols so-erp: NF=0, SKU=1, Descricao=2, Unidade=3, Qtd ERP=4, Data ERP=5, Fornecedor=6
  const headers = ['Data','Grupo','NF','SKU','Descricao','Unidade','Qtd ERP','Data ERP','Fornecedor'];
  const linhas  = [];

  document.querySelectorAll('.tab-content[id$="-so-erp"]').forEach(tab => {{
    const data  = tab.getAttribute('data-data')  || '';
    const grupo = tab.getAttribute('data-grupo') || '';
    tab.querySelectorAll('.tabela tbody tr').forEach(row => {{
      if (row.style.display === 'none') return;
      const vals = [...row.querySelectorAll('td')].map(td => td.textContent.trim());
      if (!vals[0]) return;
      linhas.push([
        data, grupo,
        vals[0],  // NF
        vals[1],  // SKU
        vals[2],  // Descricao
        vals[3],  // Unidade
        vals[4],  // Qtd ERP
        vals[5],  // Data ERP
        vals[6],  // Fornecedor
      ]);
    }});
  }});

  if (linhas.length === 0) {{ alert('Nenhum item "So no ERP" encontrado.'); return; }}

  const btn = document.getElementById('btn-exportar-soerp');
  if (btn) {{ btn.disabled = true; btn.textContent = '⏳ Gerando...'; }}
  fetch('/api/exportar_divergencias', {{
    method: 'POST',
    headers: {{'Content-Type': 'application/json'}},
    body: JSON.stringify({{headers: headers, rows: linhas}})
  }})
  .then(r => r.blob())
  .then(blob => {{
    const url = URL.createObjectURL(blob);
    const a   = document.createElement('a');
    a.href    = url;
    a.download = 'so_erp_consolidado.xlsx';
    a.click();
    URL.revokeObjectURL(url);
  }})
  .catch(e => alert('Erro ao gerar Excel: ' + e))
  .finally(() => {{
    if (btn) {{ btn.disabled = false; btn.textContent = '↙ Exportar So ERP'; }}
  }});
}}

// ── Alternar painel ──────────────────────────────────────────────────────────
function alternarView(view) {{
  document.getElementById('painel-analise').classList.toggle('ativo',  view==='analise');
  document.getElementById('painel-dashboard').classList.toggle('ativo', view==='dashboard');
  document.getElementById('vbtn-analise').classList.toggle('ativo',  view==='analise');
  document.getElementById('vbtn-dashboard').classList.toggle('ativo', view==='dashboard');
  if (view==='dashboard') inicializarDashboard();
}}

// ── Dashboard ────────────────────────────────────────────────────────────────
const _DASH = {dash_json};
const _GRUPOS_UNIQ = {grupos_json};
const _DATAS_UNIQ  = {datas_json};
const NF_LOOKUP    = {nf_lookup_json};
let _dashIniciado = false;

function inicializarDashboard() {{
  if (_dashIniciado) return;
  _dashIniciado = true;

  // Cards sumario
  let totalDiv=0, totalOk=0, totalErp=0;
  _DASH.forEach(d => {{ totalDiv += d.div; totalOk += d.ok; totalErp += d.erp; }});
  const cardsEl = document.getElementById('dash-cards-sumario');
  if (cardsEl) cardsEl.innerHTML = `
    <div class="dash-card"><div class="dash-card-label">Total Divergencias</div><div class="dash-card-valor" style="color:#f87171">${{totalDiv}}</div><div class="dash-card-sub">linhas com diff de quantidade</div></div>
    <div class="dash-card"><div class="dash-card-label">Conciliados OK</div><div class="dash-card-valor" style="color:#22c55e">${{totalOk}}</div><div class="dash-card-sub">WMS e ERP alinhados</div></div>
    <div class="dash-card"><div class="dash-card-label">So ERP</div><div class="dash-card-valor" style="color:#60a5fa">${{totalErp}}</div><div class="dash-card-sub">entradas apenas no ERP</div></div>
    <div class="dash-card"><div class="dash-card-label">Grupos</div><div class="dash-card-valor" style="color:#6366f1">${{_GRUPOS_UNIQ.length}}</div><div class="dash-card-sub">grupos analisados</div></div>
  `;

  // Tabela detalhe
  const tbody = document.getElementById('dash-tbody');
  if (tbody) {{
    tbody.innerHTML = '';
    _DASH.forEach(d => {{
      const pctBar = `<div class="pct-bar" style="width:${{Math.round(d.pct_ok)}}px;max-width:80px;background:#22c55e"></div>`;
      tbody.innerHTML += `<tr>
        <td>${{d.data}}</td><td>${{d.grupo}}</td>
        <td style="color:#22c55e">${{d.ok}}</td>
        <td style="color:#60a5fa">${{d.erp}}</td>
        <td style="color:#f87171">${{d.div}}</td>
        <td style="color:#f59e0b">${{d.div_sum}}</td>
        <td>${{d.pct_ok}}%${{pctBar}}</td>
      </tr>`;
    }});
  }}

  // Bar chart
  desenharBarChart();
}}

function desenharBarChart() {{
  const canvas = document.getElementById('dash-bar-canvas');
  if (!canvas || !_DASH.length) return;
  const dpr = window.devicePixelRatio || 1;
  const W = canvas.parentElement.clientWidth || 800;
  const H = 300;
  canvas.width  = W * dpr;
  canvas.height = H * dpr;
  canvas.style.width  = W + 'px';
  canvas.style.height = H + 'px';
  const ctx = canvas.getContext('2d');
  ctx.scale(dpr, dpr);

  const pad = {{top:20, right:20, bottom:60, left:50}};
  const cW  = W - pad.left - pad.right;
  const cH  = H - pad.top  - pad.bottom;

  // group by data, stack groups
  const byData = {{}};
  _DASH.forEach(d => {{ if (!byData[d.data]) byData[d.data] = {{}}; byData[d.data][d.grupo] = d.pct_ok; }});
  const datas = Object.keys(byData);
  const maxVal = 100;

  const nDatas  = datas.length;
  const nGrupos = _GRUPOS_UNIQ.length;
  const grpW    = cW / nDatas;
  const barW    = Math.min(grpW / (nGrupos + 1), 40);
  const cores   = ['#6366f1','#f59e0b','#22c55e','#f87171','#60a5fa'];

  // Y axis
  ctx.strokeStyle = '#333'; ctx.lineWidth = 1;
  ctx.beginPath(); ctx.moveTo(pad.left, pad.top); ctx.lineTo(pad.left, pad.top+cH); ctx.stroke();
  ctx.beginPath(); ctx.moveTo(pad.left, pad.top+cH); ctx.lineTo(pad.left+cW, pad.top+cH); ctx.stroke();
  ctx.fillStyle = '#666'; ctx.font = '10px sans-serif'; ctx.textAlign = 'right';
  for (let i=0; i<=4; i++) {{
    const y = pad.top + cH - (i/4)*cH;
    const v = Math.round((i/4)*maxVal);
    ctx.fillText(v+'%', pad.left-6, y+3);
    ctx.strokeStyle='#222'; ctx.beginPath(); ctx.moveTo(pad.left, y); ctx.lineTo(pad.left+cW, y); ctx.stroke();
  }}

  // Bars
  datas.forEach((data, di) => {{
    const baseX = pad.left + di*grpW + (grpW - nGrupos*barW) / 2;
    _GRUPOS_UNIQ.forEach((g, gi) => {{
      const val = byData[data][g] || 0;
      const bH  = val === 0 ? 0 : Math.max(2, (val/maxVal)*cH);
      const x   = baseX + gi*barW;
      const y   = pad.top + cH - bH;
      ctx.fillStyle = cores[gi % cores.length];
      ctx.fillRect(x, y, barW-2, bH);
      if (val > 0) {{
        ctx.fillStyle='#ddd'; ctx.textAlign='center'; ctx.font='10px sans-serif';
        ctx.fillText(val+'%', x + (barW-2)/2, y-4);
      }}
    }});
    // X label
    ctx.fillStyle='#888'; ctx.textAlign='center'; ctx.font='11px sans-serif';
    ctx.fillText(data, pad.left + di*grpW + grpW/2, pad.top+cH+18);
  }});

  // Legenda
  _GRUPOS_UNIQ.forEach((g, gi) => {{
    const lx = pad.left + gi*120;
    const ly = H - 14;
    ctx.fillStyle = cores[gi % cores.length];
    ctx.fillRect(lx, ly-8, 10, 10);
    ctx.fillStyle='#aaa'; ctx.textAlign='left'; ctx.font='11px sans-serif';
    ctx.fillText(g, lx+14, ly);
  }});
}}

// ── Inicializar ───────────────────────────────────────────────────────────────
{init_js}
{donuts_js}
window.addEventListener('load', carregarAjustes);

async function atualizarDados() {{
  const btn = document.getElementById('btn-atualizar');
  btn.textContent = '⏳ Atualizando...';
  btn.disabled = true;
  try {{
    const r = await fetch('/api/atualizar', {{method:'POST'}});
    if (!r.ok) throw new Error('Erro no servidor');
    const d = await r.json();
    if (d.ok) {{ location.reload(); }}
    else {{ alert('Erro ao atualizar:\\n' + d.msg); btn.textContent = '↻ Atualizar'; btn.disabled = false; }}
  }} catch(e) {{
    alert('Servidor não encontrado.\\nAbra o arquivo iniciar.bat primeiro.');
    btn.textContent = '↻ Atualizar'; btn.disabled = false;
  }}
}}
</script>
</body>
</html>"""


# ── Excel do dia ──────────────────────────────────────────────────────────────

LABELS_EXCEL = {
    "Data": "Data", "Grupo": "Cliente",
    "numero_nf": "NF", "sku": "SKU",
    "descricao_wms": "Descricao", "descricao": "Descricao",
    "unidade": "Unidade",
    "qtd_wms_usada": "Qtd WMS", "quantidade_wms": "Qtd WMS",
    "quantidade_erp": "Qtd ERP", "diff_qtd": "Diferenca",
    "valor_div": "Valor Div. (R$)",
    "peso": "Peso Liq.", "peso_wms": "Peso Liq.",
    "data_wms": "Data WMS", "data_erp": "Data ERP",
    "fornecedor": "Fornecedor",
    "Total": "Total ERP", "Conciliados OK": "Conciliados OK",
    "So no WMS": "So no WMS", "So no ERP": "So no ERP",
    "Div Quantidade": "Div. Quantidade", "% OK": "% OK",
}


FMT_BRL = '_-"R$"* #,##0.00_-;-"R$"* #,##0.00_-;_-"R$"* "-"??_-;_-@_-'
FMT_QTY = '#,##0.###'


def _formatar_sheet(ws, col_pct=None, col_num=None, col_brl=None, zebra=True, resumo=False):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Estilos base
    fill_header = PatternFill("solid", fgColor="2D3748")
    fill_par    = PatternFill("solid", fgColor="F7F8FA")
    fill_ok     = PatternFill("solid", fgColor="D1FAE5")
    fill_warn   = PatternFill("solid", fgColor="FEF3C7")
    fill_err    = PatternFill("solid", fgColor="FEE2E2")
    font_header = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
    font_body   = Font(size=10, name="Segoe UI")
    font_bold   = Font(bold=True, size=10, name="Segoe UI")
    aln_center  = Alignment(horizontal="center", vertical="center")
    aln_left    = Alignment(horizontal="left",   vertical="center")
    thin_side   = Side(style="thin", color="E2E8F0")
    border      = Border(bottom=thin_side)

    # Cabeçalho
    ws.row_dimensions[1].height = 28
    for cell in ws[1]:
        cell.fill      = fill_header
        cell.font      = font_header
        cell.alignment = aln_center

    # Freeze e auto-filtro
    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Localizar coluna % OK no Resumo
    pct_col_idx = None
    if resumo:
        for cell in ws[1]:
            if cell.value == "% OK":
                pct_col_idx = cell.column

    # Linhas de dados
    for i, row in enumerate(ws.iter_rows(min_row=2)):
        is_par = (i % 2 == 0)
        pct_val = None
        if pct_col_idx:
            pct_val = ws.cell(row=row[0].row, column=pct_col_idx).value

        # Cor de fundo por % OK (Resumo) ou zebra
        if resumo and pct_val is not None:
            if   pct_val == 100: row_fill = fill_ok
            elif pct_val >= 80:  row_fill = fill_warn
            else:                row_fill = fill_err
        elif zebra and is_par:
            row_fill = fill_par
        else:
            row_fill = None

        for cell in row:
            cell.font      = font_bold if resumo else font_body
            cell.alignment = aln_center if (col_num and cell.column in col_num) else aln_left
            cell.border    = border
            if row_fill:
                cell.fill = row_fill

            # Formatos numéricos
            if col_pct and cell.column in col_pct:
                cell.number_format = '0.0"%"'
            elif col_brl and cell.column in col_brl:
                cell.number_format = FMT_BRL
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            elif col_num and cell.column in col_num:
                v = cell.value
                is_int = isinstance(v, (int, float)) and v is not None and float(v) == int(float(v))
                cell.number_format = '#,##0' if is_int else FMT_QTY

    # Largura das colunas
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 52)


def _prep_df(df, cols):
    ex  = [c for c in cols if c in df.columns]
    out = df[ex].copy() if ex else pd.DataFrame(columns=cols)
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = out[col].dt.strftime("%d/%m/%Y").fillna("")
    out.rename(columns=LABELS_EXCEL, inplace=True)
    return out


def exportar_excel(todos: dict, caminho: str):
    from openpyxl.utils import get_column_letter

    cols_ok      = ["numero_nf","sku","descricao","unidade","qtd_wms_usada","quantidade_erp","fornecedor"]
    cols_so_wms  = ["numero_nf","sku","quantidade_wms","peso","data_wms"]
    cols_so_erp  = ["numero_nf","sku","descricao","unidade","quantidade_erp","data_erp","fornecedor"]
    cols_div_qtd = ["numero_nf","sku","descricao","unidade","qtd_wms_usada","quantidade_erp","diff_qtd","valor_div","fornecedor"]

    linhas_resumo = []
    sheets = {"Conciliados OK": [], "So no WMS": [], "So no ERP": [], "Div. Quantidade": []}

    for data_str, grupos_d in todos.items():
        for grupo, r in grupos_d.items():
            total = len(r["ok"]) + len(r["so_wms"]) + len(r["so_erp"]) + len(r["div_qtd"])
            linhas_resumo.append({
                "Data": data_str, "Grupo": grupo, "Total": total,
                "Conciliados OK": len(r["ok"]), "So no WMS": len(r["so_wms"]),
                "So no ERP": len(r["so_erp"]), "Div Quantidade": len(r["div_qtd"]),
                "% OK": round(len(r["ok"]) / total * 100, 1) if total else 0,
            })
            for df_part, cols, aba in [
                (r["ok"],      cols_ok,      "Conciliados OK"),
                (r["so_wms"],  cols_so_wms,  "So no WMS"),
                (r["so_erp"],  cols_so_erp,  "So no ERP"),
                (r["div_qtd"], cols_div_qtd, "Div. Quantidade"),
            ]:
                part = _prep_df(df_part, cols)
                part.insert(0, "Cliente", grupo)
                part.insert(0, "Data",    data_str)
                if not part.empty:
                    sheets[aba].append(part)

    df_resumo = pd.DataFrame(linhas_resumo).rename(columns=LABELS_EXCEL)

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        for aba, partes in sheets.items():
            df_aba = pd.concat(partes, ignore_index=True) if partes else pd.DataFrame()
            df_aba.to_excel(writer, sheet_name=aba, index=False)

        wb = writer.book

        # Formatar aba Resumo (com cores por % OK)
        ws_res = writer.sheets["Resumo"]
        pct_cols_res = {c.column for c in ws_res[1] if c.value == "% OK"}
        num_cols_res = {c.column for c in ws_res[1] if c.value in ("Total ERP","Conciliados OK","So no WMS","So no ERP","Div. Quantidade")}
        _formatar_sheet(ws_res, col_pct=pct_cols_res, col_num=num_cols_res, resumo=True)

        # Formatar abas de detalhe
        num_labels = {"Qtd WMS", "Qtd ERP", "Diferenca", "Peso Liq."}
        brl_labels = {"Valor Div. (R$)"}
        for aba in ("Conciliados OK", "So no WMS", "So no ERP", "Div. Quantidade"):
            if aba in writer.sheets:
                ws = writer.sheets[aba]
                num_cols = {c.column for c in ws[1] if c.value in num_labels}
                brl_cols = {c.column for c in ws[1] if c.value in brl_labels}
                _formatar_sheet(ws, col_num=num_cols, col_brl=brl_cols)


# ── Histórico acumulativo ─────────────────────────────────────────────────────

def acumular_historico(todos: dict):
    cols_ok      = ["numero_nf","sku","descricao","unidade","qtd_wms_usada","quantidade_erp","fornecedor"]
    cols_so_wms  = ["numero_nf","sku","quantidade_wms","peso","data_wms"]
    cols_so_erp  = ["numero_nf","sku","descricao","unidade","quantidade_erp","data_erp","fornecedor"]
    cols_div_qtd = ["numero_nf","sku","descricao","unidade","qtd_wms_usada","quantidade_erp","diff_qtd","valor_div","fornecedor"]

    ano = datetime.now().year

    def data_label(data_str):
        if not data_str:
            return datetime.now().strftime("%d/%m/%Y")
        try:
            d, m = data_str.split('.')
            return f"{int(d):02d}/{int(m):02d}/{ano}"
        except Exception:
            return data_str

    def com_meta(df, cols, grupo, ref):
        ex = [c for c in cols if c in df.columns]
        out = df[ex].copy() if ex else pd.DataFrame(columns=cols)
        for col in out.columns:
            if pd.api.types.is_datetime64_any_dtype(out[col]):
                out[col] = out[col].dt.strftime("%d/%m/%Y").fillna("")
        out.insert(0, "grupo",            grupo)
        out.insert(0, "data_conciliacao", ref)
        return out

    novas = {"Resumo": [], "Conciliados OK": [], "So no WMS": [], "So no ERP": [], "Div Quantidade": []}

    for data_str, grupos_d in todos.items():
        ref = data_label(data_str)
        for grupo, r in grupos_d.items():
            total_erp = len(r["ok"]) + len(r["so_erp"]) + len(r["div_qtd"])
            novas["Resumo"].append({
                "data_conciliacao": ref, "grupo": grupo,
                "total_erp": total_erp,
                "conciliados_ok": len(r["ok"]), "so_erp": len(r["so_erp"]),
                "div_quantidade": len(r["div_qtd"]),
                "pct_ok": round(len(r["ok"]) / total_erp * 100, 1) if total_erp else 0,
            })
            novas["Conciliados OK"].append(com_meta(r["ok"],      cols_ok,      grupo, ref))
            novas["So no WMS"]    .append(com_meta(r["so_wms"],   cols_so_wms,  grupo, ref))
            novas["So no ERP"]    .append(com_meta(r["so_erp"],   cols_so_erp,  grupo, ref))
            novas["Div Quantidade"].append(com_meta(r["div_qtd"], cols_div_qtd, grupo, ref))

    abas = {
        "Resumo":          pd.DataFrame(novas["Resumo"]),
        "Conciliados OK":  pd.concat(novas["Conciliados OK"],  ignore_index=True),
        "So no WMS":       pd.concat(novas["So no WMS"],       ignore_index=True),
        "So no ERP":       pd.concat(novas["So no ERP"],       ignore_index=True),
        "Div Quantidade":  pd.concat(novas["Div Quantidade"],  ignore_index=True),
    }

    if os.path.exists(HISTORICO_EXCEL):
        existente = pd.read_excel(HISTORICO_EXCEL, sheet_name=None, dtype=str)
        for aba in abas:
            if aba in existente:
                abas[aba] = pd.concat([existente[aba], abas[aba]], ignore_index=True)

    with pd.ExcelWriter(HISTORICO_EXCEL, engine="openpyxl") as writer:
        for aba, df in abas.items():
            df.to_excel(writer, sheet_name=aba, index=False)
        for ws in writer.sheets.values():
            _formatar_sheet(ws)


# ── Histórico HTML ────────────────────────────────────────────────────────────

def gerar_historico_html():
    if not os.path.exists(HISTORICO_EXCEL):
        return
    resumo = pd.read_excel(HISTORICO_EXCEL, sheet_name="Resumo", dtype=str)
    resumo["pct_ok"] = pd.to_numeric(resumo["pct_ok"], errors="coerce").fillna(0)
    grupos_hist = resumo["grupo"].unique().tolist() if "grupo" in resumo.columns else ["(todos)"]

    # Uma linha por grupo por data
    tabela_html = resumo[["data_conciliacao","grupo","total_erp","conciliados_ok","so_erp","div_quantidade","pct_ok"]].to_html(
        index=False, classes="tabela", border=0)

    datas_unicas = sorted(resumo["data_conciliacao"].unique().tolist())
    datasets_js  = []
    for g in grupos_hist:
        sub = resumo[resumo["grupo"] == g].copy()
        pcts = []
        for d in datas_unicas:
            row = sub[sub["data_conciliacao"] == d]
            pcts.append(float(row["pct_ok"].iloc[0]) if not row.empty else 0)
        datasets_js.append(f"{{label:'{g}', vals:{pcts}}}")

    datasets_str = "[" + ",".join(datasets_js) + "]"

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><title>Historico</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:'Segoe UI',sans-serif;background:#111111;color:#e0e0e0;}}
  header{{background:#1a1a1a;border-bottom:1px solid #2d2d2d;padding:18px 32px;display:flex;align-items:center;gap:16px;}}
  header h1{{color:#f1f5f9;font-size:20px;}}
  .container{{max-width:1200px;margin:0 auto;padding:28px 32px;}}
  .secao{{background:#1e1e1e;border:1px solid #2d2d2d;border-radius:12px;padding:24px;margin-bottom:24px;}}
  .secao h2{{color:#ccc;font-size:14px;font-weight:600;margin-bottom:16px;border-bottom:1px solid #2d2d2d;padding-bottom:10px;}}
  canvas{{width:100%!important;max-height:280px;}}
  .tabela{{width:100%;border-collapse:collapse;font-size:13px;}}
  .tabela th{{background:#1a1a1a;color:#94a3b8;padding:9px 12px;text-align:left;border-bottom:1px solid #333;}}
  .tabela td{{padding:8px 12px;border-bottom:1px solid #222;color:#d0d0d0;}}
  .tabela tr:hover td{{background:#222;}}
  footer{{text-align:center;color:#444;font-size:11px;padding:24px 0;}}
  a{{color:#818cf8;text-decoration:none;font-size:13px;font-weight:600;border:1px solid #6366f1;border-radius:6px;padding:6px 12px;}}
  a:hover{{background:#6366f1;color:#fff;}}
</style>
</head>
<body>
<header>
  <a href="index.html">&#8592; Inicio</a>
  <div>
    <h1>Historico de Conciliacoes</h1>
    <p style="color:#555;font-size:12px;margin-top:3px;">{len(resumo)} rodadas registradas &mdash; {len(grupos_hist)} grupos</p>
  </div>
</header>
<div class="container">
  <div class="secao">
    <h2>Evolucao do % Conciliado por Grupo</h2>
    <canvas id="grafico"></canvas>
  </div>
  <div class="secao">
    <h2>Historico detalhado</h2>
    {tabela_html}
  </div>
</div>
<footer>Historico de Conciliacoes de Estoque</footer>
<script>
const datas    = {datas_unicas};
const datasets = {datasets_str};
const cores    = ['#22c55e','#f59e0b','#60a5fa','#f87171','#8b5cf6'];
(function() {{
  const canvas = document.getElementById('grafico');
  const W = canvas.offsetWidth || 900, H = 260;
  canvas.width=W; canvas.height=H;
  const pad={{t:20,r:120,b:40,l:50}}, w=W-pad.l-pad.r, n=datas.length;
  const ctx = canvas.getContext('2d');
  ctx.fillStyle='#1e1e1e'; ctx.fillRect(0,0,W,H);
  for(let i=0;i<=5;i++) {{
    const v=i*20, y=pad.t+(H-pad.t-pad.b)*(1-v/100);
    ctx.strokeStyle='#2d2d2d';ctx.lineWidth=1;
    ctx.beginPath();ctx.moveTo(pad.l,y);ctx.lineTo(W-pad.r,y);ctx.stroke();
    ctx.fillStyle='#555';ctx.font='10px Segoe UI';ctx.textAlign='right';
    ctx.fillText(v+'%',pad.l-6,y+3);
  }}
  datasets.forEach((ds,di) => {{
    const cor=cores[di%cores.length];
    ctx.beginPath();ctx.strokeStyle=cor;ctx.lineWidth=2;
    ds.vals.forEach((v,i) => {{
      const x=pad.l+(n===1?w/2:(w/(n-1))*i);
      const y=pad.t+(H-pad.t-pad.b)*(1-v/100);
      i===0?ctx.moveTo(x,y):ctx.lineTo(x,y);
    }});
    ctx.stroke();
    ds.vals.forEach((v,i) => {{
      const x=pad.l+(n===1?w/2:(w/(n-1))*i);
      const y=pad.t+(H-pad.t-pad.b)*(1-v/100);
      ctx.beginPath();ctx.arc(x,y,3,0,Math.PI*2);ctx.fillStyle=cor;ctx.fill();
    }});
    ctx.fillStyle=cor;ctx.font='11px Segoe UI';ctx.textAlign='left';
    const lastX=pad.l+(n===1?w/2:w), lastY=pad.t+(H-pad.t-pad.b)*(1-ds.vals[n-1]/100);
    ctx.fillText(ds.label,lastX+8,lastY+4);
  }});
  ctx.fillStyle='#777';ctx.font='11px Segoe UI';ctx.textAlign='center';
  datas.forEach((d,i) => {{
    const x=pad.l+(n===1?w/2:(w/(n-1))*i);
    ctx.fillText(d,x,H-pad.b+16);
  }});
}})();
</script>
</body></html>"""

    with open(HISTORICO_HTML, "w", encoding="utf-8") as f:
        f.write(html)


# ── Index ─────────────────────────────────────────────────────────────────────

def gerar_index(arquivos_html: list):
    if not os.path.exists(HISTORICO_EXCEL):
        return
    resumo  = pd.read_excel(HISTORICO_EXCEL, sheet_name="Resumo", dtype=str)
    resumo["pct_ok"] = pd.to_numeric(resumo["pct_ok"], errors="coerce").fillna(0)

    # Agrupar por arquivo (data+hora) → cada HTML pode ter múltiplos grupos
    linhas = ""
    for html_file in reversed(arquivos_html):
        ts_str = html_file.replace("conciliacao_", "").replace(".html", "")
        try:
            ts = datetime.strptime(ts_str, "%Y%m%d_%H%M%S")
            data_label = ts.strftime("%d/%m/%Y %H:%M")
        except Exception:
            data_label = ts_str

        excel_nome = html_file.replace(".html", ".xlsx")
        html_link  = f'<a href="{html_file}" style="color:#818cf8;">&#9654; Abrir</a>'
        excel_link = f'<a href="{excel_nome}" style="color:#3b82f6;">&#8595; Excel</a>' if os.path.exists(os.path.join(OUTPUT_DIR, excel_nome)) else "—"
        linhas += f"<tr><td>{data_label}</td><td>{html_link}</td><td>{excel_link}</td></tr>\n"

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8"><title>Conciliacao de Estoque</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:'Segoe UI',sans-serif;background:#111111;color:#e0e0e0;}}
  header{{background:#1a1a1a;border-bottom:1px solid #2d2d2d;padding:20px 32px;display:flex;justify-content:space-between;align-items:center;}}
  header h1{{color:#f1f5f9;font-size:20px;font-weight:600;}}
  header p{{color:#555;font-size:12px;margin-top:4px;}}
  .container{{max-width:1100px;margin:0 auto;padding:28px 32px;}}
  .atalhos{{display:flex;gap:14px;margin-bottom:28px;flex-wrap:wrap;}}
  .card-atalho{{flex:1;min-width:180px;background:#1e1e1e;border:1px solid #2d2d2d;border-radius:10px;padding:16px 20px;text-decoration:none;color:inherit;transition:border-color .2s;display:flex;align-items:center;gap:12px;}}
  .card-atalho:hover{{border-color:#6366f1;}}
  .card-atalho .icone{{font-size:22px;}}
  .card-atalho .titulo{{font-size:13px;font-weight:600;color:#ccc;}}
  .card-atalho .sub{{font-size:11px;color:#555;margin-top:2px;}}
  .secao{{background:#1e1e1e;border:1px solid #2d2d2d;border-radius:12px;padding:24px;}}
  .secao-topo{{display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;padding-bottom:10px;border-bottom:1px solid #2d2d2d;}}
  .secao-topo h2{{font-size:14px;font-weight:600;color:#aaa;}}
  .filtro{{padding:7px 12px;background:#141414;border:1px solid #333;border-radius:6px;color:#ccc;font-size:12px;outline:none;width:220px;}}
  .filtro:focus{{border-color:#6366f1;}}
  table{{width:100%;border-collapse:collapse;font-size:13px;}}
  th{{background:#1a1a1a;color:#94a3b8;padding:9px 14px;text-align:left;border-bottom:1px solid #333;}}
  td{{padding:9px 14px;border-bottom:1px solid #222;}}
  tr:hover td{{background:#222;}}
  a{{text-decoration:none;}}
  footer{{text-align:center;color:#444;font-size:11px;padding:28px 0;}}
</style>
</head>
<body>
<header>
  <div>
    <a href="index.html" style="font-size:11px;color:#555;text-decoration:none;display:inline-block;margin-bottom:6px;">&#8592; Home</a>
    <h1>Conciliacao de Entrada</h1><p>Historico de rodadas</p>
  </div>
  <div style="font-size:12px;color:#555;">{len(arquivos_html)} rodadas registradas</div>
</header>
<div class="container">
  <div class="atalhos">
    <a class="card-atalho" href="{arquivos_html[-1] if arquivos_html else '#'}">
      <div class="icone">&#9654;</div>
      <div><div class="titulo">Ultima Conciliacao</div><div class="sub">Abrir painel mais recente</div></div>
    </a>
  </div>
  <div class="secao">
    <div class="secao-topo">
      <h2>Todas as rodadas</h2>
      <input class="filtro" type="text" placeholder="Filtrar por data..." onkeyup="filtrar(this)">
    </div>
    <table id="tb">
      <thead><tr><th>Data / Hora</th><th>Dashboard</th><th>Excel</th></tr></thead>
      <tbody>{linhas}</tbody>
    </table>
  </div>
</div>
<footer>Conciliacao de Estoque &mdash; atualizado automaticamente</footer>
<script>
function filtrar(input) {{
  const t=input.value.toLowerCase();
  document.querySelectorAll('#tb tbody tr').forEach(tr=>{{
    tr.style.display=tr.textContent.toLowerCase().includes(t)?'':'none';
  }});
}}
</script>
</body></html>"""

    with open(INDEX_HTML, "w", encoding="utf-8") as f:
        f.write(html)


def gerar_home():
    """Gera a página inicial com cards de navegação para cada módulo."""
    html = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Conciliacao de Estoque</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: 'Segoe UI', system-ui, sans-serif;
    background: #0a0a0a;
    color: #e2e8f0;
    min-height: 100vh;
    display: flex;
    flex-direction: column;
  }
  header {
    padding: 32px 48px 0;
    border-bottom: 1px solid #1a1a1a;
    padding-bottom: 28px;
  }
  header h1 { font-size: 22px; font-weight: 700; color: #f1f5f9; letter-spacing: -.3px; }
  header p  { font-size: 13px; color: #555; margin-top: 4px; }
  .container {
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    padding: 60px 24px;
    gap: 32px;
  }
  .titulo-modulos {
    font-size: 13px;
    color: #444;
    text-transform: uppercase;
    letter-spacing: 1.5px;
    font-weight: 600;
  }
  .cards {
    display: flex;
    gap: 28px;
    flex-wrap: wrap;
    justify-content: center;
  }
  .card {
    width: 280px;
    background: #111;
    border: 1px solid #222;
    border-radius: 16px;
    padding: 36px 28px;
    text-decoration: none;
    color: inherit;
    display: flex;
    flex-direction: column;
    gap: 14px;
    transition: border-color .2s, transform .15s, background .2s;
    position: relative;
    overflow: hidden;
  }
  .card:hover { border-color: #444; background: #161616; transform: translateY(-3px); }
  .card.disabled { cursor: default; opacity: .5; pointer-events: none; }
  .card-icone {
    width: 52px; height: 52px;
    border-radius: 12px;
    display: flex; align-items: center; justify-content: center;
    font-size: 24px;
  }
  .card.entrada .card-icone { background: #0d2a1a; }
  .card.saida   .card-icone { background: #1a0d2a; }
  .card-titulo {
    font-size: 17px;
    font-weight: 700;
    color: #f1f5f9;
    line-height: 1.3;
  }
  .card-desc {
    font-size: 12px;
    color: #555;
    line-height: 1.6;
  }
  .card-badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 10px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: .8px;
    align-self: flex-start;
  }
  .card.entrada .card-badge { background: #0d2a1a; color: #22c55e; }
  .card.saida   .card-badge { background: #1a1a2a; color: #6366f1; }
  .card-arrow {
    position: absolute;
    bottom: 28px; right: 28px;
    font-size: 18px;
    color: #333;
    transition: color .2s, right .15s;
  }
  .card:hover .card-arrow { color: #666; right: 24px; }
  footer {
    text-align: center;
    color: #333;
    font-size: 11px;
    padding: 24px 0 32px;
  }
</style>
</head>
<body>
<header>
  <h1>Conciliacao de Estoque</h1>
  <p>Selecione o modulo desejado</p>
</header>
<div class="container">
  <div class="titulo-modulos">Modulos disponíveis</div>
  <div class="cards">

    <a class="card entrada" href="conciliacao_latest.html">
      <div class="card-icone">&#128229;</div>
      <div class="card-titulo">Conciliacao de Entrada</div>
      <div class="card-desc">
        Reconciliacao WMS vs ERP para notas fiscais de entrada.
        Detecta divergencias de quantidade, itens ausentes e ajustes pendentes.
      </div>
      <span class="card-badge">Ativo</span>
      <span class="card-arrow">&#8594;</span>
    </a>

    <a class="card saida" href="saida_conciliacao_latest.html">
      <div class="card-icone">&#128228;</div>
      <div class="card-titulo">Conciliacao de Saida</div>
      <div class="card-desc">
        Reconciliacao de notas fiscais de saída entre WMS e ERP.
      </div>
      <span class="card-badge">Ativo</span>
      <span class="card-arrow">&#8594;</span>
    </a>

  </div>
</div>
<footer>COMFRIO FOOD SERVICE &mdash; Conciliacao de Estoque</footer>
</body>
</html>"""
    with open(HOME_HTML, "w", encoding="utf-8") as f:
        f.write(html)


# ── Main ──────────────────────────────────────────────────────────────────────

def _rodar_grupo(grupo, wms_path, erp_path):
    """Carrega, filtra e concilia WMS vs ERP, separando por data do WMS.
    Retorna {data_str: resultados}, onde data_str é 'DD.MM' ou '' se sem data."""
    print(f"  WMS: {os.path.basename(wms_path)}")
    df_wms = carregar_arquivo(wms_path, WMS_COLUNAS)
    print(f"  {len(df_wms)} registros WMS")
    print(f"  ERP: {os.path.basename(erp_path)}")
    df_erp = carregar_arquivo(erp_path, ERP_COLUNAS)
    print(f"  {len(df_erp)} registros ERP (bruto)")

    if "Tipo Docto." in df_erp.columns:
        antes = len(df_erp)
        df_erp = df_erp[df_erp["Tipo Docto."].str.strip().str.upper() == "N"].copy()
        print(f"  {antes - len(df_erp)} registros ERP removidos (Tipo Docto. != N)")

    padrao = ERP_SKU_PADRAO.get(grupo)
    if padrao:
        mask      = df_erp["sku"].str.match(padrao, na=False)
        ignorados = (~mask).sum()
        df_erp    = df_erp[mask].copy()
        if ignorados:
            print(f"  {ignorados} SKUs ignorados (fora do padrao '{padrao}')")
    print(f"  {len(df_erp)} registros ERP (validos)")

    # Datas únicas no WMS (DATA INCLUSÃO)
    datas_wms = sorted(df_wms["data"].dropna().dt.date.unique())

    if not datas_wms:
        r = conciliar(df_wms, df_erp)
        print(f"  OK:{len(r['ok'])}  WMS:{len(r['so_wms'])}  ERP:{len(r['so_erp'])}  Div:{len(r['div_qtd'])}")
        return {"": r}

    # Mapeia cada NF do WMS para a data em que foi recebida
    nf_para_data = {}
    for dt in datas_wms:
        for nf in df_wms.loc[df_wms["data"].dt.date == dt, "numero_nf"].dropna().unique():
            nf_para_data.setdefault(nf, dt)

    # Atribui cada linha do ERP à data WMS correspondente:
    # 1° prioridade: NF encontrada no WMS naquela data
    # 2° prioridade: data ERP mais próxima da data WMS
    def _data_grupo_erp(row):
        nf = row["numero_nf"]
        if nf in nf_para_data:
            return nf_para_data[nf]
        erp_dt = row["data"]
        if pd.isna(erp_dt):
            return datas_wms[0]
        erp_date = erp_dt.date() if hasattr(erp_dt, "date") else erp_dt
        return min(datas_wms, key=lambda d: abs((d - erp_date).days))

    df_erp = df_erp.copy()
    df_erp["_data_grupo"] = df_erp.apply(_data_grupo_erp, axis=1)

    resultados = {}
    for dt in datas_wms:
        data_str = f"{dt.day:02d}.{dt.month:02d}"
        wms_d = df_wms[df_wms["data"].dt.date == dt].copy()
        erp_d = df_erp[df_erp["_data_grupo"] == dt].drop(columns=["_data_grupo"]).copy()
        r = conciliar(wms_d, erp_d)
        print(f"  [{data_str}] OK:{len(r['ok'])}  WMS:{len(r['so_wms'])}  ERP:{len(r['so_erp'])}  Div:{len(r['div_qtd'])}")
        resultados[data_str] = r

    return resultados


def main():
    if len(sys.argv) >= 3:
        wms_path = sys.argv[1]
        erp_path = sys.argv[2]
        grupo    = sys.argv[3] if len(sys.argv) > 3 else os.path.basename(wms_path)
        pares    = [(grupo, "", wms_path, erp_path)]
    else:
        pares = detectar_pares(DADOS_DIR)
        if not pares:
            print(f"Nenhum par WMS/ERP encontrado em '{DADOS_DIR}/'.")
            sys.exit(1)
        print(f"Pares detectados: {[g for g, _, _, _ in pares]}")

    # Rodar cada grupo e coletar resultados separados por data
    todos = {}  # {data_str: {grupo: resultados}}
    for grupo, _data_arquivo, wms_path, erp_path in pares:
        print(f"\n[{grupo}]")
        por_data = _rodar_grupo(grupo, wms_path, erp_path)
        for data_str, r in por_data.items():
            todos.setdefault(data_str, {})[grupo] = r

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    datas_str = " | ".join(d for d in sorted(todos) if d) or datetime.now().strftime("%d/%m/%Y")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    saida_html  = os.path.join(OUTPUT_DIR, f"conciliacao_{timestamp}.html")
    saida_excel = os.path.join(OUTPUT_DIR, f"conciliacao_{timestamp}.xlsx")

    html = gerar_dashboard_multi(todos, datas_str, [], [])
    with open(saida_html, "w", encoding="utf-8") as f:
        f.write(html)
    latest_html = os.path.join(OUTPUT_DIR, "conciliacao_latest.html")
    with open(latest_html, "w", encoding="utf-8") as f:
        f.write(html)

    exportar_excel(todos, saida_excel)
    acumular_historico(todos)
    gerar_historico_html()

    gerar_home()

    print(f"\nDashboard: {saida_html}")
    print(f"Excel:     {saida_excel}")
    print(f"Historico: {HISTORICO_EXCEL}")
    print(f"Home:      {HOME_HTML}")


if __name__ == "__main__":
    main()
